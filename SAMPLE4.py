# recitation_vosk_enhanced_v2.py
# Enhanced version with improved accuracy mechanisms
# Requirements: pip install vosk pyaudio openpyxl rapidfuzz numpy scikit-learn librosa

import os
import sys
import threading
import numpy as np
import pyaudio
import time
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from vosk import Model, KaldiRecognizer
import json
from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz, process
import atexit
import traceback
import re

# Optional imports for pronunciation training
try:
    import librosa
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    import pickle
    PRONUNCIATION_TRAINING_AVAILABLE = True
    print("✅ Pronunciation training libraries loaded")
except ImportError as e:
    PRONUNCIATION_TRAINING_AVAILABLE = False
    print(f"⚠️ Pronunciation training disabled. Missing: {e}")

# ------------------ Settings ------------------
OUTPUT_XLSX = "recitation_log.xlsx"
LOG_TXT = "recitation_log.txt"
SHEET_NAME = "Recitations"

ROSTER = [
    "Abellera", "Alba", "Belleza", "Bindoy", "Borja", "Cipriano", "Cruz", "Dasco", 
    "Dela Cruz", "Dela Peña", "Dela Rama", "Deris", "Dublois", "Espiritu", "Fatalla", 
    "Franco", "Garcia", "Guevarra", "Ison", "Jonsay", "Lagdan", "Llamas", "Loyola", 
    "Marquez", "Martinez", "Mijares", "Narag", "Negru", "Nieva", "Olid", "Ordonez", 
    "Pacheco", "Paculan", "Pandi", "Paz", "Peña", "Requiño", "Sajo", "Samonte",
    "Saribay", "Serviano", "Silvestre", "Surell", "Torres", "Blanco, Jerard", 
    "Blanco, Keon", "Blanco, Raizjhea", "Ocol, Sean", "Okol, Xyzea"
]

# Enhanced keyword configuration with phonetic variations
def generate_phonetic_variations(name):
    """Generate common mispronunciations for Filipino names."""
    variations = [name.lower()]
    
    # Common substitutions in Filipino speech recognition
    substitutions = {
        'ñ': ['n', 'ny'],
        'j': ['h', 'dy'],
        'z': ['s'],
        'v': ['b'],
        'f': ['p'],
        'x': ['ks', 's'],
    }
    
    for orig, replacements in substitutions.items():
        for repl in replacements:
            variations.append(name.lower().replace(orig, repl))
    
    return list(set(variations))

# Build enhanced keyword list
keywords = []
for name in ROSTER:
    keywords.append({"keyword": name.lower(), "boost": 50.0})
    # Add variations with lower boost
    for variation in generate_phonetic_variations(name)[1:]:
        keywords.append({"keyword": variation, "boost": 20.0})

# Save keywords for Vosk
try:
    with open("kws.json", "w", encoding="utf-8") as f:
        json.dump(keywords, f, ensure_ascii=False, indent=2)
    KWS_LIST = [k["keyword"] for k in keywords]
except Exception:
    KWS_LIST = [name.lower() for name in ROSTER]

# Improved matching thresholds
NAME_MATCH_THRESHOLD = 65  # Increased from 50
CONFIDENCE_THRESHOLD = 0.6  # Minimum confidence for Vosk results
SAMPLE_RATE = 16000
CHANNELS = 1
CHUNK = 1024
BUFFER_SECONDS = 4  # Increased from 3
SILENCE_THRESHOLD = 600  # Increased from 500
SILENCE_DURATION = 1.0  # Increased from 0.8

# Model paths
FILIPINO_MODEL_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python313\vosk-model-tl-ph-generic-0.6"
ENGLISH_MODEL_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python313\vosk-model-en-us-0.22"
DATASET_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python312\dataset"
AUDIO_DIR = os.path.join(DATASET_PATH, "audio")
TRANSCRIPTIONS_DIR = os.path.join(DATASET_PATH, "transcriptions")
PRONUNCIATION_MODEL_PATH = "pronunciation_model.pkl"

USE_FILIPINO = True
VOSK_MODEL_PATH = FILIPINO_MODEL_PATH if USE_FILIPINO else ENGLISH_MODEL_PATH

# Common misrecognition patterns
COMMON_MISRECOGNITIONS = {
    'balangkas': ['abellera', 'belleza'],
    'alba': ['alba', 'alva'],
    'belize': ['belleza'],
    'bindoy': ['bindoy', 'mendoy'],
    'fine': ['five'],
    'for': ['four'],
    'sex': ['six'],
    'to': ['two'],
}


class PronunciationTrainer:
    """Handles custom pronunciation training."""
    
    def __init__(self):
        self.pronunciation_data = {}
        self.trained = False
        
        if PRONUNCIATION_TRAINING_AVAILABLE:
            try:
                self.text_vectorizer = TfidfVectorizer(analyzer='char', ngram_range=(2, 4))
                self.load_or_create_model()
            except Exception as e:
                print(f"❌ Pronunciation trainer init error: {e}")
        else:
            print("⚠️ Using standard matching only")
    
    def load_pronunciation_data(self):
        """Load pronunciation data from dataset folder."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return
            
        if not os.path.exists(AUDIO_DIR):
            print(f"⚠️ Audio directory not found: {AUDIO_DIR}")
            return
            
        if not os.path.exists(TRANSCRIPTIONS_DIR):
            print(f"⚠️ Transcriptions directory not found: {TRANSCRIPTIONS_DIR}")
            return
        
        print("📚 Loading pronunciation training data...")
        loaded_count = 0
        
        try:
            for audio_file in os.listdir(AUDIO_DIR):
                if not audio_file.lower().endswith(('.wav', '.mp3', '.flac')):
                    continue
                    
                base_name = os.path.splitext(audio_file)[0]
                transcript_path = os.path.join(TRANSCRIPTIONS_DIR, f"{base_name}.txt")
                
                if not os.path.exists(transcript_path):
                    continue
                
                try:
                    with open(transcript_path, 'r', encoding='utf-8') as f:
                        transcript = f.read().strip().lower()
                    
                    matched_name = None
                    for name in ROSTER:
                        if name.lower() in transcript:
                            matched_name = name
                            break
                    
                    if matched_name:
                        if matched_name not in self.pronunciation_data:
                            self.pronunciation_data[matched_name] = {'transcripts': []}
                        
                        self.pronunciation_data[matched_name]['transcripts'].append(transcript)
                        loaded_count += 1
                        print(f"✅ Loaded: {matched_name} -> {transcript}")
                        
                except Exception as e:
                    print(f"⚠️ Error processing {audio_file}: {e}")
            
            print(f"📊 Loaded {loaded_count} pronunciation samples")
            
            if self.pronunciation_data:
                self.train_text_similarity()
                
        except Exception as e:
            print(f"❌ Error loading pronunciation data: {e}")
    
    def train_text_similarity(self):
        """Train text similarity model."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return
            
        try:
            all_texts = []
            
            for name, data in self.pronunciation_data.items():
                all_texts.extend(data['transcripts'])
            
            if all_texts:
                all_texts.extend([name.lower() for name in ROSTER])
                self.text_vectorizer.fit(all_texts)
                self.trained = True
                print("✅ Text similarity model trained")
            
        except Exception as e:
            print(f"❌ Error training similarity model: {e}")
    
    def save_model(self):
        """Save trained model."""
        if not PRONUNCIATION_TRAINING_AVAILABLE or not self.trained:
            return
            
        try:
            model_data = {
                'pronunciation_data': self.pronunciation_data,
                'text_vectorizer': self.text_vectorizer,
                'trained': self.trained
            }
            
            with open(PRONUNCIATION_MODEL_PATH, 'wb') as f:
                pickle.dump(model_data, f)
            
            print(f"✅ Model saved: {PRONUNCIATION_MODEL_PATH}")
            
        except Exception as e:
            print(f"❌ Error saving model: {e}")
    
    def load_model(self):
        """Load saved model."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return False
            
        try:
            if os.path.exists(PRONUNCIATION_MODEL_PATH):
                with open(PRONUNCIATION_MODEL_PATH, 'rb') as f:
                    model_data = pickle.load(f)
                
                self.pronunciation_data = model_data.get('pronunciation_data', {})
                self.text_vectorizer = model_data.get('text_vectorizer', 
                    TfidfVectorizer(analyzer='char', ngram_range=(2, 4)))
                self.trained = model_data.get('trained', False)
                
                print(f"✅ Model loaded: {PRONUNCIATION_MODEL_PATH}")
                return True
                
        except Exception as e:
            print(f"❌ Error loading model: {e}")
        
        return False
    
    def load_or_create_model(self):
        """Load existing model or create new one."""
        if not self.load_model():
            print("🔧 Creating new pronunciation model...")
            self.load_pronunciation_data()
            if self.trained:
                self.save_model()
    
    def match_pronunciation(self, recognized_text):
        """Enhanced name matching using training data."""
        if not PRONUNCIATION_TRAINING_AVAILABLE or not self.trained:
            return None, 0
        
        if not recognized_text.strip():
            return None, 0
        
        recognized_text = recognized_text.lower().strip()
        best_match = None
        best_score = 0
        
        try:
            recognized_vector = self.text_vectorizer.transform([recognized_text])
            
            for name, data in self.pronunciation_data.items():
                for transcript in data['transcripts']:
                    transcript_vector = self.text_vectorizer.transform([transcript])
                    similarity = cosine_similarity(recognized_vector, transcript_vector)[0][0]
                    score = similarity * 100
                    
                    if score > best_score and score >= (NAME_MATCH_THRESHOLD * 0.8):
                        best_score = score
                        best_match = name
            
            return best_match, best_score
            
        except Exception as e:
            print(f"❌ Error in pronunciation matching: {e}")
            return None, 0


def save_text_log():
    """Save a summary of the recognition session when app exits."""
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary = [
            "\n========== SESSION SUMMARY ==========",
            f"Session End: {ts}",
        ]

        if 'app' in globals() and isinstance(app, RecitationSystem):
            try:
                accuracy = (app.successful_matches / app.total_attempts * 100) if app.total_attempts > 0 else 0
            except Exception:
                accuracy = 0.0
            avg_speed = (sum(app.processing_times) / len(app.processing_times)) if app.processing_times else 0.0
            summary.extend([
                f"Total Attempts: {app.total_attempts}",
                f"Successful Matches: {app.successful_matches}",
                f"Accuracy: {accuracy:.1f}%",
                f"Average Recognition Speed: {avg_speed:.2f}s",
            ])

            if getattr(app, "recognition_log", None):
                summary.append("\nRecent recognitions (most recent first):")
                for entry in app.recognition_log[-30:][::-1]:
                    summary.append(entry)

        summary.append("=====================================\n")

        with open(LOG_TXT, "a", encoding="utf-8") as f:
            f.write("\n".join(summary) + "\n")

    except Exception as e:
        print(f"⚠️ Could not save session log: {e}")


class RecitationSystem:
    def __init__(self):
        print("🚀 Initializing Enhanced Recitation System v2...")
        
        self.pronunciation_trainer = PronunciationTrainer()
        self.setup_excel()
        self.setup_vosk()
        self.setup_audio()
        
        self.audio_buffer = []
        self.last_speech_time = time.time()
        self.processing_lock = threading.Lock()
        self.listening = False

        self.total_attempts = 0
        self.successful_matches = 0
        self.processing_times = []
        self.recognition_log = []
        self.last_recognized_name = None  # Track last recognized name
        self.confirmation_mode = False
        
        self.setup_gui()
        print("✅ System initialized successfully!")

    def setup_excel(self):
        """Setup Excel file."""
        try:
            if not os.path.exists(OUTPUT_XLSX):
                wb = Workbook()
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(["Name", "Points", "Timestamp", "Total"])
                wb.save(OUTPUT_XLSX)

            self.wb = load_workbook(OUTPUT_XLSX)
            if SHEET_NAME in self.wb.sheetnames:
                self.ws = self.wb[SHEET_NAME]
            else:
                self.ws = self.wb.create_sheet(SHEET_NAME)
                self.ws.append(["Name", "Points", "Timestamp", "Total"])

            headers = [cell.value for cell in self.ws[1]]
            if len(headers) < 4 or headers[3] != "Total":
                self.ws.cell(row=1, column=4, value="Total")
                self.wb.save(OUTPUT_XLSX)
            
            print("✅ Excel setup complete")
            
        except Exception as e:
            print(f"❌ Excel setup error: {e}")

    def setup_vosk(self):
        """Setup Vosk model with enhanced configuration."""
        try:
            print(f"Loading Vosk model: {VOSK_MODEL_PATH}")
            start_time = time.time()
            
            if not os.path.exists(VOSK_MODEL_PATH):
                print(f"❌ Model not found: {VOSK_MODEL_PATH}")
                self.model = None
                self.recognizer = None
                self.model_info = "Model: NOT FOUND"
                return
        
            self.model = Model(VOSK_MODEL_PATH)
            # Enable word-level confidence scores
            self.recognizer = KaldiRecognizer(self.model, SAMPLE_RATE)
            self.recognizer.SetWords(True)  # Changed to True for confidence scores
        
            load_time = time.time() - start_time
            print(f"✅ Vosk model loaded in {load_time:.1f}s")
            
            training_status = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
            self.model_info = f"Vosk | {'Filipino' if USE_FILIPINO else 'English'} | {training_status}"
            
        except Exception as e:
            print(f"❌ Vosk error: {e}")
            self.model = None
            self.recognizer = None
            self.model_info = "Model: LOAD FAILED"

    def setup_audio(self):
        """Setup audio stream."""
        try:
            self.p = pyaudio.PyAudio()
            self.stream = self.p.open(
                format=pyaudio.paInt16,
                channels=CHANNELS,
                rate=SAMPLE_RATE,
                input=True,
                frames_per_buffer=CHUNK
            )
            print("✅ Audio setup complete")
            
        except Exception as e:
            print(f"❌ Audio setup error: {e}")
            self.stream = None

    def setup_gui(self):
        """Setup GUI interface."""
        try:
            self.root = tk.Tk()
            self.root.title("🎤 Enhanced Recitation System v2")
            self.root.geometry("900x800")
            self.root.configure(bg='#f0f0f0')

            # Title
            title = tk.Label(self.root, text="🎤 ENHANCED RECITATION SYSTEM v2",
                           font=('Arial', 16, 'bold'), bg='#f0f0f0', fg='#2c3e50')
            title.pack(pady=10)

            # Model info
            self.model_label = tk.Label(self.root, text=getattr(self, 'model_info', 'Loading...'),
                                       font=('Arial', 9), bg='#f0f0f0', fg='#7f8c8d')
            self.model_label.pack()

            # Training status
            if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
                sample_count = sum(len(data['transcripts']) for data in self.pronunciation_trainer.pronunciation_data.values())
                training_info = f"🧠 Pronunciation Training: ACTIVE ({sample_count} samples)"
                training_color = '#27ae60'
            else:
                training_info = "⚠️ Pronunciation Training: INACTIVE"
                training_color = '#e67e22'
            
            self.training_label = tk.Label(self.root, text=training_info,
                                          font=('Arial', 9, 'bold'), bg='#f0f0f0', fg=training_color)
            self.training_label.pack()

            # Accuracy tips
            tips_frame = tk.Frame(self.root, bg='#ecf0f1', relief='ridge', bd=2)
            tips_frame.pack(pady=5, padx=20, fill='x')
            
            tips_label = tk.Label(tips_frame, 
                                 text="💡 Tips: Speak clearly | Pause after name | Say full name then points",
                                 font=('Arial', 9), bg='#ecf0f1', fg='#34495e')
            tips_label.pack(pady=5)

            # Status
            self.status_label = tk.Label(self.root, text="Ready to start",
                                        font=('Arial', 12), bg='#f0f0f0', fg='#27ae60')
            self.status_label.pack(pady=5)

            # Stats
            stats_frame = tk.Frame(self.root, bg='#f0f0f0')
            stats_frame.pack(pady=5)

            self.accuracy_label = tk.Label(stats_frame, text="Accuracy: 0% (0/0)",
                                          font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#e74c3c')
            self.accuracy_label.pack(side='left', padx=10)

            self.speed_label = tk.Label(stats_frame, text="Avg Speed: 0.0s",
                                       font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#3498db')
            self.speed_label.pack(side='left', padx=10)

            # Speech Recognition
            speech_frame = ttk.LabelFrame(self.root, text="🎙️ Speech Recognition", padding=10)
            speech_frame.pack(pady=10, padx=20, fill='x')

            self.speech_button = tk.Button(speech_frame, text="START LISTENING",
                                          command=self.toggle_listening,
                                          font=('Arial', 12, 'bold'),
                                          bg='#3498db', fg='white', width=20)
            self.speech_button.pack(pady=5)

            self.transcription_label = tk.Label(speech_frame, text="Say: 'Full Name + Points' (e.g., 'Belleza five')",
                                               font=('Arial', 10), wraplength=700)
            self.transcription_label.pack(pady=5)
            
            # Confidence indicator
            self.confidence_label = tk.Label(speech_frame, text="Confidence: --",
                                            font=('Arial', 9), fg='#7f8c8d')
            self.confidence_label.pack(pady=2)

            # Manual Entry
            manual_frame = ttk.LabelFrame(self.root, text="⌨️ Manual Entry", padding=10)
            manual_frame.pack(pady=10, padx=20, fill='x')

            tk.Label(manual_frame, text="Student:").grid(row=0, column=0, sticky='w', pady=2)
            self.student_var = tk.StringVar()
            self.student_combo = ttk.Combobox(manual_frame, textvariable=self.student_var,
                                             values=ROSTER, width=18, state='readonly')
            self.student_combo.grid(row=0, column=1, padx=5, pady=2)

            tk.Label(manual_frame, text="Points:").grid(row=0, column=2, sticky='w', padx=(20, 0), pady=2)
            self.points_var = tk.StringVar()
            self.points_entry = tk.Entry(manual_frame, textvariable=self.points_var, width=10)
            self.points_entry.grid(row=0, column=3, padx=5, pady=2)

            submit_btn = tk.Button(manual_frame, text="ADD POINTS", command=self.manual_submit,
                                  font=('Arial', 10, 'bold'), bg='#27ae60', fg='white')
            submit_btn.grid(row=0, column=4, padx=10, pady=2)

            quick_frame = tk.Frame(manual_frame)
            quick_frame.grid(row=1, column=0, columnspan=5, pady=10)

            tk.Label(quick_frame, text="Quick Add:").pack(side='left')
            for points in [1, 2, 3, 5]:
                btn = tk.Button(quick_frame, text=f"+{points}",
                               command=lambda p=points: self.quick_add_points(p),
                               font=('Arial', 9), bg='#f39c12', fg='white', width=5)
                btn.pack(side='left', padx=2)

            # Scores Table
            scores_frame = ttk.LabelFrame(self.root, text="📊 Current Scores", padding=10)
            scores_frame.pack(pady=10, padx=20, fill='both', expand=True)

            self.tree = ttk.Treeview(scores_frame, columns=('Name', 'Latest', 'Total'), show='headings', height=6)
            self.tree.heading('Name', text='Student Name')
            self.tree.heading('Latest', text='Latest Points')
            self.tree.heading('Total', text='Total Points')

            self.tree.column('Name', width=260)
            self.tree.column('Latest', width=120, anchor='center')
            self.tree.column('Total', width=120, anchor='center')

            scrollbar = ttk.Scrollbar(scores_frame, orient='vertical', command=self.tree.yview)
            self.tree.configure(yscrollcommand=scrollbar.set)

            self.tree.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')

            # Log
            log_frame = ttk.LabelFrame(self.root, text="📋 Recognition Log", padding=5)
            log_frame.pack(pady=10, padx=20, fill='both', expand=True)

            log_controls = tk.Frame(log_frame)
            log_controls.pack(fill='x', pady=(0, 5))

            tk.Label(log_controls, text="Speech recognition activity:",
                    font=('Arial', 9, 'bold')).pack(side='left')

            clear_btn = tk.Button(log_controls, text="Clear Log", command=self.clear_log,
                                 font=('Arial', 8), bg='#95a5a6', fg='white', width=10)
            clear_btn.pack(side='right')

            log_text_frame = tk.Frame(log_frame)
            log_text_frame.pack(fill='both', expand=True)

            self.log_text = tk.Text(log_text_frame, height=8, font=('Consolas', 9), wrap='word',
                                   bg='#2c3e50', fg='#ecf0f1', insertbackground='white')
            log_scrollbar = ttk.Scrollbar(log_text_frame, orient='vertical', command=self.log_text.yview)
            self.log_text.configure(yscrollcommand=log_scrollbar.set)

            self.log_text.pack(side='left', fill='both', expand=True)
            log_scrollbar.pack(side='right', fill='y')

            self.add_to_log("=== ENHANCED AI RECITATION SYSTEM v2 READY ===")
            self.add_to_log("🎯 Improved accuracy with multi-stage matching")
            if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
                sample_count = sum(len(data['transcripts']) for data in self.pronunciation_trainer.pronunciation_data.values())
                self.add_to_log(f"🧠 Enhanced matching with {sample_count} pronunciation samples")
            self.add_to_log("Click 'START LISTENING' to begin recognition...")

            self.points_entry.bind('<Return>', lambda e: self.manual_submit())

            self.refresh_scores()
            print("✅ GUI setup complete")
            
        except Exception as e:
            print(f"❌ GUI setup error: {e}")
            traceback.print_exc()

    def match_name_multi_stage(self, recognized_text, confidence=1.0):
        """Multi-stage name matching with improved accuracy."""
        if not recognized_text.strip():
            return None, 0, "empty"
        
        recognized_text = recognized_text.lower().strip()
        
        # Stage 1: Check if confidence is too low
        if confidence < CONFIDENCE_THRESHOLD:
            self.add_to_log(f"⚠️ Low confidence ({confidence:.2f}), skipping match")
            return None, 0, "low_confidence"
        
        # Stage 2: Try pronunciation training first (if available)
        if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
            name, score = self.pronunciation_trainer.match_pronunciation(recognized_text)
            if name and score >= NAME_MATCH_THRESHOLD:
                return name, score, "pronunciation_training"
        
        # Stage 3: Direct exact match (case-insensitive)
        for roster_name in ROSTER:
            if roster_name.lower() == recognized_text:
                return roster_name, 100, "exact_match"
        
        # Stage 4: Check for known misrecognitions
        for misheard, correct_names in COMMON_MISRECOGNITIONS.items():
            if misheard in recognized_text:
                for correct in correct_names:
                    if correct.lower() in [n.lower() for n in ROSTER]:
                        roster_name = next(n for n in ROSTER if n.lower() == correct.lower())
                        return roster_name, 85, "misrecognition_correction"
        
        # Stage 5: Word-by-word fuzzy matching with multiple algorithms
        words = recognized_text.split()
        best_match = None
        best_score = 0
        best_word = None
        
        for word in words:
            # Skip very short words and number words
            if len(word) < 3 or word in ['one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten',
                                          'isa', 'dalawa', 'tatlo', 'apat', 'lima']:
                continue
            
            # Try multiple fuzzy matching algorithms
            roster_lower = [n.lower() for n in ROSTER]
            
            # Token sort ratio (good for word order changes)
            match1, score1, _ = process.extractOne(word, roster_lower, scorer=fuzz.token_sort_ratio)
            
            # Partial ratio (good for partial matches)
            match2, score2, _ = process.extractOne(word, roster_lower, scorer=fuzz.partial_ratio)
            
            # WRatio (weighted ratio - balanced)
            match3, score3, _ = process.extractOne(word, roster_lower, scorer=fuzz.WRatio)
            
            # Take the best score among all methods
            scores = [(match1, score1), (match2, score2), (match3, score3)]
            local_best = max(scores, key=lambda x: x[1])
            
            if local_best[1] > best_score:
                best_score = local_best[1]
                best_match = next(n for n in ROSTER if n.lower() == local_best[0])
                best_word = word
        
        # Stage 6: Validate the match
        if best_match and best_score >= NAME_MATCH_THRESHOLD:
            # Extra validation: check if match length is similar to recognized word
            if best_word:
                length_diff = abs(len(best_match) - len(best_word))
                if length_diff > 5:  # Too different in length
                    best_score *= 0.8  # Penalize
                    
                # Check if starts with same letter
                if best_match[0].lower() == best_word[0].lower():
                    best_score = min(100, best_score * 1.1)  # Boost
            
            if best_score >= NAME_MATCH_THRESHOLD:
                return best_match, best_score, "fuzzy_match"
        
        return None, 0, "no_match"

    def toggle_listening(self):
        """Toggle speech recognition."""
        if not self.listening:
            if self.model and self.stream:
                self.listening = True
                self.speech_button.config(text="STOP LISTENING", bg='#e74c3c')
                self.status_label.config(text="🎤 AI is listening...", fg='#e74c3c')
                self.add_to_log("🎤 STARTED listening for speech")
                threading.Thread(target=self.audio_loop, daemon=True).start()
            else:
                messagebox.showerror("Error", "Speech recognition not available.")
                self.add_to_log("❌ Speech recognition unavailable")
        else:
            self.listening = False
            self.speech_button.config(text="START LISTENING", bg='#3498db')
            self.status_label.config(text="Stopped listening", fg='#f39c12')
            self.add_to_log("⏹️ STOPPED listening")

    def audio_loop(self):
        """Main audio processing loop."""
        while self.listening and self.stream:
            try:
                data = self.stream.read(CHUNK, exception_on_overflow=False)
                current_time = time.time()

                if not self.is_silence(data):
                    self.audio_buffer.append(data)
                    self.last_speech_time = current_time
                else:
                    if self.audio_buffer and (current_time - self.last_speech_time) > SILENCE_DURATION:
                        self.process_audio_buffer()

                max_chunks = int(SAMPLE_RATE * BUFFER_SECONDS / CHUNK)
                if len(self.audio_buffer) > max_chunks:
                    self.process_audio_buffer()

            except Exception as e:
                print(f"Audio loop error: {e}")
                self.root.after(0, lambda: self.add_to_log(f"[ERROR] Audio error: {e}"))
                break

    def is_silence(self, data: bytes) -> bool:
        """Check if audio data is silence."""
        try:
            audio_array = np.frombuffer(data, dtype=np.int16)
            return np.abs(audio_array).mean() < SILENCE_THRESHOLD
        except Exception:
            return False

    def process_audio_buffer(self):
        """Process accumulated audio with enhanced confidence checking."""
        with self.processing_lock:
            if not self.audio_buffer or not self.model:
                self.audio_buffer = []
                return

            try:
                start_time = time.time()
                self.root.after(0, lambda: self.status_label.config(text="🔄 Processing...", fg='#f39c12'))

                audio_data = b''.join(self.audio_buffer)

                # Create fresh recognizer with word-level confidence
                recognizer = KaldiRecognizer(self.model, SAMPLE_RATE)
                recognizer.SetWords(True)

                # Process with Vosk
                accepted = recognizer.AcceptWaveform(audio_data)
                if accepted:
                    result_text = recognizer.Result()
                else:
                    result_text = recognizer.FinalResult()

                # Parse result with confidence
                try:
                    result_json = json.loads(result_text)
                    text = result_json.get("text", "").strip()
                    
                    # Calculate average confidence from word-level scores
                    confidence = 1.0
                    if "result" in result_json:
                        word_confs = [w.get("conf", 1.0) for w in result_json["result"]]
                        if word_confs:
                            confidence = sum(word_confs) / len(word_confs)
                    
                except Exception:
                    text = ""
                    confidence = 0.0

                processing_time = time.time() - start_time
                self.processing_times.append(processing_time)
                timestamp = datetime.now().strftime("%H:%M:%S")

                if text:
                    self.total_attempts += 1
                    self.root.after(0, lambda t=text: self.transcription_label.config(text=f"Heard: '{t}'"))
                    self.root.after(0, lambda c=confidence: self.confidence_label.config(
                        text=f"Confidence: {c:.2%}",
                        fg='#27ae60' if c >= 0.8 else '#f39c12' if c >= 0.6 else '#e74c3c'
                    ))

                    method = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
                    log_entry = f"[{timestamp}] {method}: '{text}' | Conf: {confidence:.2%} | {processing_time:.2f}s"
                    self.root.after(0, lambda: self.add_to_log(log_entry))

                    success = self.process_speech_text(text, processing_time, confidence)
                    if success:
                        self.successful_matches += 1
                        self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] ✅ SUCCESS - Points logged"))
                    else:
                        self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] ❌ FAILED - No valid match"))

                    self.update_performance_stats()
                else:
                    self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] ❌ No speech detected ({processing_time:.2f}s)"))
                    self.root.after(0, lambda: self.status_label.config(text="🎤 Listening...", fg='#27ae60'))

                self.audio_buffer = []

            except Exception as e:
                print(f"Processing error: {e}")
                self.root.after(0, lambda: self.add_to_log(f"[ERROR] Processing failed: {str(e)}"))
                self.audio_buffer = []

    def process_speech_text(self, text, processing_time=0.0, confidence=1.0):
        """Process recognized speech text with improved validation."""
        # Extract points first
        points = self.extract_points(text)
        if not points:
            self.root.after(0, lambda: self.status_label.config(
                text="❌ No points detected. Say name and points clearly.", fg='#e74c3c'))
            return False

        # Match name with multi-stage approach
        name, score, method = self.match_name_multi_stage(text, confidence)
        
        if name:
            # Log detailed match info
            match_info = f"Match: {name} | Score: {score:.1f}% | Method: {method} | Confidence: {confidence:.2%}"
            self.root.after(0, lambda: self.add_to_log(f"🎯 {match_info}"))
            
            # Extra confirmation for low-confidence matches
            if score < 75 or confidence < 0.7:
                self.root.after(0, lambda n=name, p=points, s=score: self.status_label.config(
                    text=f"⚠️ Low confidence match: {n} +{p} pts ({s:.0f}%) - Use manual if incorrect", 
                    fg='#f39c12'))
            else:
                self.root.after(0, lambda n=name, p=points, s=score: self.status_label.config(
                    text=f"✅ {n} +{p} pts! (Match: {s:.0f}%, Conf: {confidence:.0%})", 
                    fg='#27ae60'))
            
            self.log_points(name, points, processing_time, match_score=score, method=method, confidence=confidence)
            return True
        else:
            # Provide helpful feedback
            suggestion = self.suggest_closest_match(text)
            if suggestion:
                self.root.after(0, lambda s=suggestion: self.status_label.config(
                    text=f"❌ No match. Did you mean: {s}? Use manual entry.", fg='#e74c3c'))
            else:
                self.root.after(0, lambda: self.status_label.config(
                    text="❌ No name match. Speak clearly or use manual entry.", fg='#e74c3c'))
            return False

    def suggest_closest_match(self, text):
        """Suggest the closest matching name for user guidance."""
        words = text.lower().split()
        best_suggestion = None
        best_score = 0
        
        for word in words:
            if len(word) < 3:
                continue
            match, score, _ = process.extractOne(word, [n.lower() for n in ROSTER])
            if score > best_score and score >= 40:  # Lower threshold for suggestions
                best_score = score
                best_suggestion = next(n for n in ROSTER if n.lower() == match)
        
        return best_suggestion if best_score >= 40 else None

    def extract_points(self, text):
        """Extract points from text with improved recognition."""
        # Enhanced number word mappings
        number_words = {
            # English
            "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
            "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
            # Filipino
            "isa": 1, "dalawa": 2, "tatlo": 3, "apat": 4, "lima": 5,
            "anim": 6, "pito": 7, "walo": 8, "siyam": 9, "sampu": 10,
            # Common misrecognitions
            "won": 1, "to": 2, "too": 2, "tree": 3, "for": 4, "fore": 4,
            "fine": 5, "sex": 6, "ate": 8, "tin": 10
        }

        # First try to find digits
        digits = re.findall(r'\d+', text)
        if digits:
            try:
                return int(digits[0])
            except Exception:
                pass

        # Then try number words
        words = text.lower().split()
        for word in words:
            clean_word = ''.join(c for c in word if c.isalpha())
            if clean_word in number_words:
                return number_words[clean_word]

        return None

    def manual_submit(self):
        """Handle manual point submission."""
        student = self.student_var.get()
        points_str = self.points_var.get()

        if not student:
            messagebox.showwarning("Warning", "Please select a student")
            return

        try:
            points = int(points_str)
            if points <= 0:
                raise ValueError()
        except Exception:
            messagebox.showwarning("Warning", "Please enter a valid positive number for points")
            return

        self.log_points(student, points, processing_time=0.0, manual=True)
        self.points_var.set("")
        self.status_label.config(text=f"✅ {student} +{points} points (manual)", fg='#27ae60')

    def quick_add_points(self, points):
        """Quick add points to selected student."""
        student = self.student_var.get()
        if not student:
            messagebox.showwarning("Warning", "Please select a student first")
            return

        self.log_points(student, points, processing_time=0.0, manual=True)
        self.status_label.config(text=f"✅ {student} +{points} points (quick)", fg='#27ae60')

    def log_points(self, name, points, processing_time=0.0, match_score=None, manual=False, method="Standard", confidence=1.0):
        """Log points to Excel and text log."""
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        found = False

        for row in range(2, self.ws.max_row + 1):
            cell_name = self.ws.cell(row=row, column=1).value
            if cell_name and cell_name.lower() == name.lower():
                old_points = self.ws.cell(row=row, column=2).value
                if old_points:
                    self.ws.cell(row=row, column=2, value=f"{old_points}+{points}")
                else:
                    self.ws.cell(row=row, column=2, value=str(points))

                self.ws.cell(row=row, column=3, value=ts)

                current_total = self.ws.cell(row=row, column=4).value or 0
                self.ws.cell(row=row, column=4, value=current_total + points)

                found = True
                break

        if not found:
            next_row = self.ws.max_row + 1
            self.ws.cell(row=next_row, column=1, value=name)
            self.ws.cell(row=next_row, column=2, value=str(points))
            self.ws.cell(row=next_row, column=3, value=ts)
            self.ws.cell(row=next_row, column=4, value=points)

        try:
            self.wb.save(OUTPUT_XLSX)
        except Exception as e:
            print(f"Failed to save Excel: {e}")

        total_now = self.get_total(name)
        speed_part = f" | Speed: {processing_time:.2f}s" if processing_time > 0 else ""
        score_part = f" | {method}Score: {match_score:.0f}%" if match_score is not None else ""
        conf_part = f" | Conf: {confidence:.0%}" if confidence < 1.0 else ""
        manual_part = " (manual)" if manual else ""
        text_log_line = f"{ts} | {name} +{points}{manual_part} | Total: {total_now}{score_part}{conf_part}{speed_part}"

        try:
            with open(LOG_TXT, "a", encoding="utf-8") as f:
                f.write(text_log_line + "\n")
        except Exception as e:
            print(f"Failed to write text log: {e}")

        self.recognition_log.append(text_log_line)
        self.root.after(0, self.refresh_scores)

    def get_total(self, name):
        """Return numeric total for a student."""
        for row in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=1).value
            if val and val.lower() == name.lower():
                return self.ws.cell(row=row, column=4).value or 0
        return 0

    def update_performance_stats(self):
        """Update performance statistics display."""
        if self.total_attempts > 0:
            accuracy = (self.successful_matches / self.total_attempts) * 100
            accuracy_color = '#27ae60' if accuracy >= 70 else '#f39c12' if accuracy >= 50 else '#e74c3c'
            self.accuracy_label.config(
                text=f"Accuracy: {accuracy:.0f}% ({self.successful_matches}/{self.total_attempts})",
                fg=accuracy_color
            )
        else:
            self.accuracy_label.config(text="Accuracy: 0% (0/0)", fg='#e74c3c')

        if self.processing_times:
            avg_speed = sum(self.processing_times) / len(self.processing_times)
            speed_color = '#27ae60' if avg_speed <= 2.0 else '#f39c12' if avg_speed <= 4.0 else '#e74c3c'
            self.speed_label.config(text=f"Avg Speed: {avg_speed:.2f}s", fg=speed_color)
        else:
            self.speed_label.config(text="Avg Speed: 0.00s", fg='#3498db')

    def add_to_log(self, message):
        """Add message to recognition log widget."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"

        self.log_text.insert(tk.END, formatted_message + "\n")
        self.log_text.see(tk.END)

        try:
            if "✅ SUCCESS" in message:
                self.log_text.tag_add("success", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("success", foreground='#2ecc71')
            elif "❌" in message or "ERROR" in message:
                self.log_text.tag_add("error", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("error", foreground='#e74c3c')
            elif "🎯" in message:
                self.log_text.tag_add("match", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("match", foreground='#9b59b6')
            elif "Enhanced:" in message:
                self.log_text.tag_add("enhanced", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("enhanced", foreground='#3498db')
        except Exception:
            pass

        try:
            lines = int(self.log_text.index('end-1c').split('.')[0])
            if lines > 200:
                self.log_text.delete("1.0", f"{lines-200}.0")
        except Exception:
            pass

    def clear_log(self):
        """Clear the recognition log and reset stats."""
        self.log_text.delete("1.0", tk.END)
        self.total_attempts = 0
        self.successful_matches = 0
        self.processing_times = []
        self.recognition_log = []
        self.update_performance_stats()
        self.add_to_log("=== LOG CLEARED - STATS RESET ===")
        status = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
        self.add_to_log(f"System ready with {status} matching")

    def refresh_scores(self):
        """Refresh the Treeview with current Excel scores."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in range(2, self.ws.max_row + 1):
            name = self.ws.cell(row=row, column=1).value
            history = self.ws.cell(row=row, column=2).value or ""
            total = self.ws.cell(row=row, column=4).value or 0

            latest = 0
            if isinstance(history, str) and history.strip():
                try:
                    last_part = history.split("+")[-1]
                    latest = int(last_part)
                except Exception:
                    try:
                        latest = int(history)
                    except Exception:
                        latest = 0
            elif isinstance(history, (int, float)):
                latest = int(history)

            if name:
                self.tree.insert('', 'end', values=(name, latest, total))

    def run(self):
        """Start the GUI mainloop."""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def on_closing(self):
        """Cleanup on window close."""
        self.listening = False
        try:
            if self.stream:
                self.stream.stop_stream()
                self.stream.close()
            if hasattr(self, 'p'):
                self.p.terminate()
        except Exception:
            pass

        try:
            self.wb.save(OUTPUT_XLSX)
        except Exception:
            pass

        save_text_log()
        try:
            self.root.destroy()
        except Exception:
            pass


if __name__ == "__main__":
    print("🚀 Starting Enhanced Recitation System v2...")
    try:
        app = RecitationSystem()
        atexit.register(save_text_log)
        app.run()
    except Exception as e:
        print(f"❌ Failed to start application: {e}")
        traceback.print_exc()
        input("Press Enter to exit...")
    print("👋 Application closed.")
