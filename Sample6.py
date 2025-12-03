import os
import sys
import threading
import numpy as np
import pyaudio
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from vosk import Model, KaldiRecognizer
import json
from openpyxl import Workbook, load_workbook
from rapidfuzz import process, fuzz
from rapidfuzz.fuzz import token_set_ratio, partial_ratio
import atexit
import traceback
import wave
import tempfile
from collections import defaultdict
import re
import unicodedata

# Optional imports for pronunciation training
try:
    import librosa
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    import pickle
    PRONUNCIATION_TRAINING_AVAILABLE = True
    print("✅ Pronunciation training libraries loaded")
except ImportError as e:
    PRONUNCIATION_TRAINING_AVAILABLE = False
    print(f"⚠️ Pronunciation training disabled. Missing: {e}")

# Optional imports for speaker diarization
try:
    from pyannote.audio import Pipeline
    from pyannote.audio.pipelines.speaker_verification import PretrainedSpeakerEmbedding
    import torch
    import torchaudio
    SPEAKER_DIARIZATION_AVAILABLE = True
    print("✅ Speaker diarization libraries loaded")
except ImportError as e:
    SPEAKER_DIARIZATION_AVAILABLE = False
    print(f"⚠️ Speaker diarization disabled. Missing: {e}")

# SETTINGS FOR RECITATION SYSTEM
OUTPUT_XLSX = "recitation_log.xlsx"
LOG_TXT = "recitation_log.txt"
SHEET_NAME = "Recitations"

# SEPARATE ROSTERS FOR DIFFERENT MATCHING STRATEGIES

# Roster 1: Students with UNIQUE last names (last name only needed)
UNIQUE_LASTNAME_ROSTER = [
    "Abellera", "Alba", "Belleza", "Bindoy", "Borja", "Cipriano", "Cruz", "Dasco", "Dela Cruz", "Dela Peña", "Dela Rama", "Deris", "Dublois",
    "Espiritu", "Fatalla", "Franco", "Garcia", "Guevarra", "Ison", "Jonsay", "Lagdan", "Llamas", "Loyola", "Marquez", "Martinez", "Mijares",
    "Narag", "Negru", "Nieva", "Olid", "Ordonez", "Pacheco", "Paculan", "Pandi", "Paz", "Peña", "Requiño", "Sajo", "Samonte",
    "Saribay", "Serviano", "Silvestre", "Surell", "Torres"
]

# Roster 2: Students with DUPLICATE last names (need first + last name)
DUPLICATE_LASTNAME_ROSTER = [
    "Blanco, Jerard", "Blanco, Keon", "Blanco, Raizjhea", "Ocol, Sean", "Okol, Xyzea"
]

# Combined roster for total count
ROSTER = UNIQUE_LASTNAME_ROSTER + DUPLICATE_LASTNAME_ROSTER

def normalize_text(s: str) -> str:
    """Improved text normalization with better compound name handling."""
    if not s:
        return ""
    # Normalize unicode characters
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    
    # Standardize compound names more flexibly
    s = re.sub(r'\b(de\s+la|dela)\s+', 'dela ', s)
    
    # Clean up spaces and standardize
    s = re.sub(r"\s+", " ", s)
    s = s.strip()
    
    return s

def expand_name_variants(name: str) -> list:
    base = normalize_text(name)
    variants = {base}
    if " de la " in base:
        variants.add(base.replace(" de la ", " dela "))
    if " dela " in base:
        variants.add(base.replace(" dela ", " de la "))
    return list(variants)

# Add to create_dual_roster_keywords function
def create_dual_roster_keywords():
    keywords = []
    
    # Add name variations and context
    for lastname in UNIQUE_LASTNAME_ROSTER:
        # Basic keyword with higher boost
        keywords.append({"keyword": normalize_text(lastname), "boost": 40.0})
        
        # Add variations for compound names
        if 'dela' in lastname.lower() or 'de la' in lastname.lower():
            base = normalize_text(lastname)
            variants = [
                base.replace('dela', 'de la'),
                base.replace('de la', 'dela'),
                base.split()[-1]  # Last part only
            ]
            for variant in variants:
                keywords.append({"keyword": variant, "boost": 35.0})
    
    # Special handling for problematic names
    special_names = {
        "Loyola": ["loyola", "loyola"],
        "Belleza": ["belleza", "bellesa", "beleza"],
        "Jonsay": ["jonsay", "honsay", "konsay"],
        "Requiño": ["requino", "rekino", "requiño"],
        "Blanco": ["blanco", "blanko"],
        "Ocol": ["ocol", "okol"]
    }
    
    for base_name, variants in special_names.items():
        for variant in variants:
            keywords.append({"keyword": variant, "boost": 45.0})
    
    # Handle duplicate names (Blanco, Ocol) with first names
    duplicate_variations = {
        "Blanco, Jerard": ["blanco jerard", "jerard blanco", "blanco, jerard"],
        "Blanco, Keon": ["blanco keon", "keon blanco", "blanco, keon"],
        "Blanco, Raizjhea": ["blanco raizjhea", "raizjhea blanco", "blanco, raizjhea"],
        "Ocol, Sean": ["ocol sean", "sean ocol", "ocol, sean"],
        "Ocol, Xyzea": ["ocol xyzea", "xyzea ocol", "ocol, xyzea"]
    }
    
    for full_name, variants in duplicate_variations.items():
        for variant in variants:
            keywords.append({"keyword": variant, "boost": 50.0})
    
    return keywords  # ADD THIS LINE - IT WAS MISSING!

# NOW CALL THE FUNCTION AND SAVE - Replace lines 63-69 with:
keywords = create_dual_roster_keywords()  # CREATE the keywords

# Save keywords for Vosk
try:
    with open("kws_recitation.json", "w", encoding="utf-8") as f:
        json.dump(keywords, f, ensure_ascii=False, indent=2)
    KWS_LIST = [k["keyword"] for k in keywords]
except Exception:
    KWS_LIST = []

# Enhanced matching settings - More balanced thresholds
NAME_MATCH_THRESHOLD = 70  # Return to original value
FULL_NAME_MATCH_THRESHOLD = 85  # Return to original value
# Commit thresholds
UNIQUE_COMMIT_THRESHOLD = 80  # Return to original value
FULLNAME_COMMIT_THRESHOLD = 85  # Return to original value

# Audio settings
SAMPLE_RATE = 16000
CHANNELS = 1
CHUNK = 1024
BUFFER_SECONDS = 4
SILENCE_THRESHOLD = 200 # Adjusted from 500 for better silence detection
SILENCE_DURATION = 0.5 # Adjusted from 1.0 for more responsive segmentation

# Speaker recognition settings
TEACHER_VOICE_MODEL_PATH = "teacher_voice_model.pkl"
SPEAKER_SIMILARITY_THRESHOLD = 0.55  # Lowered from 0.75 for less strict matching
MIN_SPEECH_DURATION = 0.5

# Model paths
FILIPINO_MODEL_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python313\vosk-model-tl-ph-generic-0.6"
ENGLISH_MODEL_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python313\vosk-model-en-us-0.22"
DATASET_PATH = r"C:\Users\HP\AppData\Local\Programs\Python\Python312\dataset"
TEACHER_VOICE_DIR = os.path.join(DATASET_PATH, "teacher_voice")
AUDIO_DIR = os.path.join(DATASET_PATH, "audio")
TRANSCRIPTIONS_DIR = os.path.join(DATASET_PATH, "transcripts")
PRONUNCIATION_MODEL_PATH = "pronunciation_model_recitation.pkl"

USE_FILIPINO = True
VOSK_MODEL_PATH = FILIPINO_MODEL_PATH if USE_FILIPINO else ENGLISH_MODEL_PATH

class PronunciationTrainer:
    """Enhanced pronunciation trainer with audio-based learning."""
    
    def __init__(self):
        self.pronunciation_data = {}
        self.audio_features = {}
        self.trained = False
        
        if PRONUNCIATION_TRAINING_AVAILABLE:
            try:
                self.text_vectorizer = TfidfVectorizer(analyzer='char', ngram_range=(2, 4))
                self.load_or_create_model()
            except Exception as e:
                print(f"❌ Pronunciation trainer init error: {e}")
        else:
            print("⚠️ Using standard matching only")
    
    def load_pronunciation_data(self):
        """Load pronunciation data from audio recordings and transcriptions."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return
        
        if not os.path.exists(AUDIO_DIR) or not os.path.exists(TRANSCRIPTIONS_DIR):
            print(f"⚠️ Dataset directories not found")
            return
        
        print("📚 Loading pronunciation training data from recordings...")
        loaded_count = 0
        
        try:
            # Get all audio files
            audio_files = [f for f in os.listdir(AUDIO_DIR) 
                          if f.lower().endswith(('.wav', '.mp3', '.flac'))]
            
            for audio_file in audio_files:
                base_name = os.path.splitext(audio_file)[0]
                transcript_path = os.path.join(TRANSCRIPTIONS_DIR, f"{base_name}.txt")
                audio_path = os.path.join(AUDIO_DIR, audio_file)
                
                if not os.path.exists(transcript_path):
                    print(f"⚠️ Missing transcript for {audio_file}")
                    continue
                
                try:
                    # Read the transcript
                    with open(transcript_path, 'r', encoding='utf-8') as f:
                        transcript = f.read().strip().lower()
                    
                    # Match transcript to student name
                    matched_name = self.match_transcript_to_student(transcript)
                    
                    if matched_name:
                        # Extract audio features
                        audio_features = self.extract_audio_features(audio_path)
                        
                        if audio_features is not None:
                            if matched_name not in self.pronunciation_data:
                                self.pronunciation_data[matched_name] = {
                                    'transcripts': [],
                                    'audio_features': []
                                }
                            
                            self.pronunciation_data[matched_name]['transcripts'].append(transcript)
                            self.pronunciation_data[matched_name]['audio_features'].append(audio_features)
                            loaded_count += 1
                            print(f"✅ Loaded: {matched_name} -> '{transcript}'")
                        
                except Exception as e:
                    print(f"⚠️ Error processing {audio_file}: {e}")
            
            print(f"📊 Loaded {loaded_count} pronunciation samples")
            self.train_text_similarity()
            
        except Exception as e:
            print(f"❌ Error loading pronunciation data: {e}")

    def match_number_pronunciation(self, text):
        """Match number pronunciations using enhanced patterns."""
        if not text:
            return None, 0
            
        text_norm = normalize_text(text)
        
        # Enhanced number patterns including Filipino
        number_patterns = {
            r'\b(one|isa|uno)\b': (1, 95),
            r'\b(two|dalawa|dos)\b': (2, 95),
            r'\b(three|tatlo|tres)\b': (3, 95),
            r'\b(four|apat|cuatro)\b': (4, 95),
            r'\b(five|lima|cinco)\b': (5, 95),
            r'\b([1-9]|10)\b': (None, 100),  # Extract digit
            r'\bplus\s*([1-9]|10)\b': (None, 90),
            r'\b([1-9]|10)\s*points?\b': (None, 95),
        }
        
        best_number = None
        best_score = 0
        
        for pattern, (number, score) in number_patterns.items():
            match = re.search(pattern, text_norm, re.IGNORECASE)
            if match:
                if number is None:
                    try:
                        extracted = int(match.group(1) if len(match.groups()) > 0 else match.group(0))
                        if 1 <= extracted <= 10:
                            number = extracted
                    except:
                        continue
                
                if number and score > best_score:
                    best_number = number
                    best_score = score
        
        return best_number, best_score

    def match_transcript_to_student(self, transcript):
        """Match transcript text to student name."""
        transcript_norm = normalize_text(transcript)
        
        # Try exact matches first
        for name in ROSTER:
            name_norm = normalize_text(name)
            if name_norm in transcript_norm:
                return name
        
        # Try fuzzy matching
        best_match = None
        best_score = 0
        
        for name in ROSTER:
            score = fuzz.token_set_ratio(transcript_norm, normalize_text(name))
            if score > best_score and score >= 60:  # Lower threshold for transcript matching
                best_score = score
                best_match = name
        
        return best_match
    
    def extract_audio_features(self, audio_path):
        """Extract audio features for pronunciation matching."""
        try:
            import librosa
            
            # Load audio
            y, sr = librosa.load(audio_path, sr=SAMPLE_RATE)
            
            if len(y) < sr * 0.5:  # Less than 0.5 seconds
                return None
            
            # Extract features
            features = {}
            
            # MFCC features
            mfcc = librosa.feature.mfcc(y=y, sr=sr, n_mfcc=13)
            features['mfcc_mean'] = np.mean(mfcc, axis=1)
            features['mfcc_std'] = np.std(mfcc, axis=1)
            
            # Spectral features
            spectral_centroids = librosa.feature.spectral_centroid(y=y, sr=sr)
            features['spectral_centroid'] = np.mean(spectral_centroids)
            
            # Zero crossing rate
            zcr = librosa.feature.zero_crossing_rate(y)
            features['zcr'] = np.mean(zcr)
            
            # Chroma features
            chroma = librosa.feature.chroma_stft(y=y, sr=sr)
            features['chroma_mean'] = np.mean(chroma, axis=1)
            
            return features
            
        except Exception as e:
            print(f"❌ Error extracting audio features from {audio_path}: {e}")
            return None
    
    def train_text_similarity(self):
        """Train text similarity model with collected data."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return
            
        try:
            all_texts = []
            
            # Add all transcripts
            for name, data in self.pronunciation_data.items():
                all_texts.extend(data['transcripts'])
            
            # Add standard name references
            all_texts.extend([normalize_text(name) for name in ROSTER])
            
            if all_texts:
                self.text_vectorizer.fit(all_texts)
                self.trained = True
                print("✅ Text similarity model trained with audio-based data")
            
        except Exception as e:
            print(f"❌ Error training similarity model: {e}")
    
    def match_pronunciation(self, recognized_text):
        """Enhanced pronunciation matching using audio training data."""
        if not PRONUNCIATION_TRAINING_AVAILABLE or not self.trained:
            return None, 0
        
        if not recognized_text.strip():
            return None, 0
        
        recognized_text = normalize_text(recognized_text)
        best_match = None
        best_score = 0
        
        try:
            recognized_vector = self.text_vectorizer.transform([recognized_text])
            
            # Check against trained pronunciation data
            for name, data in self.pronunciation_data.items():
                for transcript in data['transcripts']:
                    transcript_vector = self.text_vectorizer.transform([transcript])
                    similarity = cosine_similarity(recognized_vector, transcript_vector)[0][0]
                    score = similarity * 100
                    
                    if score > best_score and score >= 50:  # Lowered threshold for audio-trained data
                        best_score = score
                        best_match = name
            
            return best_match, best_score
            
        except Exception as e:
            print(f"❌ Error in pronunciation matching: {e}")
            return None, 0
    
    def save_model(self):
        """Save trained pronunciation model."""
        if not PRONUNCIATION_TRAINING_AVAILABLE or not self.trained:
            return
            
        try:
            model_data = {
                'pronunciation_data': self.pronunciation_data,
                'text_vectorizer': self.text_vectorizer,
                'trained': self.trained,
                'version': '3.0'  # Updated version for audio features
            }
            
            with open(PRONUNCIATION_MODEL_PATH, 'wb') as f:
                pickle.dump(model_data, f)
            
            print(f"✅ Enhanced pronunciation model saved")
            
        except Exception as e:
            print(f"❌ Error saving model: {e}")
    
    def load_model(self):
        """Load saved pronunciation model."""
        if not PRONUNCIATION_TRAINING_AVAILABLE:
            return False
            
        try:
            if os.path.exists(PRONUNCIATION_MODEL_PATH):
                with open(PRONUNCIATION_MODEL_PATH, 'rb') as f:
                    model_data = pickle.load(f)
                
                self.pronunciation_data = model_data.get('pronunciation_data', {})
                self.text_vectorizer = model_data.get('text_vectorizer', 
                    TfidfVectorizer(analyzer='char', ngram_range=(2, 4)))
                self.trained = model_data.get('trained', False)
                
                if self.pronunciation_data:
                    sample_count = sum(len(data['transcripts']) for data in self.pronunciation_data.values())
                    print(f"✅ Loaded pronunciation model with {sample_count} audio samples")
                    return True
                
        except Exception as e:
            print(f"❌ Error loading pronunciation model: {e}")
        
        return False
    
    def load_or_create_model(self):
        """Load existing model or create new one from audio data."""
        if not self.load_model():
            print("🔧 Creating new pronunciation model from audio recordings...")
            self.load_pronunciation_data()
            if self.trained:
                self.save_model()
class DualRosterAnalyzer:
    """Analyzes dual roster structure and creates appropriate mappings."""
    
    def __init__(self, unique_roster, duplicate_roster):
        self.unique_roster = unique_roster
        self.duplicate_roster = duplicate_roster
        self.full_roster = unique_roster + duplicate_roster
        self.name_mappings = self.create_name_mappings()
        
        print(f"📊 Dual Roster Analysis:")
        print(f"   Unique surnames: {len(unique_roster)} students")
        print(f"   Duplicate surnames: {len(duplicate_roster)} students") 
        print(f"   Total: {len(self.full_roster)} students")
        
        # Show unique surnames
        if len(unique_roster) <= 10:
            print(f"   Unique: {', '.join(unique_roster)}")
        else:
            print(f"   Unique: {', '.join(unique_roster[:10])}... (+{len(unique_roster)-10} more)")
        
        # Show duplicate surname groups
        duplicate_groups = defaultdict(list)
        for full_name in duplicate_roster:
            if ',' in full_name:
                lastname = full_name.split(',')[0].strip()
                duplicate_groups[lastname].append(full_name)
        
        print(f"   Duplicate surname groups: {len(duplicate_groups)}")
        for lastname, students in list(duplicate_groups.items())[:5]:
            print(f"     {lastname}: {len(students)} students")
    
    def create_name_mappings(self):
        """Create mappings for both roster types."""
        mappings = {
            'unique_lastnames': {},      # "cruz" -> "Cruz"
            'duplicate_fullnames': {},   # "garcia maria" -> "Garcia, Maria"
            'firstname_lastname': {},    # "maria garcia" -> "Garcia, Maria"
            'all_students': {}
        }
        
        # Map unique surnames (normalized keys)
        for lastname in self.unique_roster:
            key = normalize_text(lastname)
            mappings['unique_lastnames'][key] = lastname
            mappings['all_students'][key] = lastname

        # Map duplicate surname full names (normalized keys)
        for full_name in self.duplicate_roster:
            if ',' in full_name:
                last_name, first_name = full_name.split(',', 1)
                last_name = last_name.strip()
                first_name = first_name.strip()

                full_key = normalize_text(f"{last_name} {first_name}")
                reversed_key = normalize_text(f"{first_name} {last_name}")

                mappings['duplicate_fullnames'][full_key] = full_name
                mappings['firstname_lastname'][reversed_key] = full_name
                mappings['all_students'][full_key] = full_name
                mappings['all_students'][reversed_key] = full_name
        
        return mappings
    
    def is_unique_lastname(self, lastname):
        """Check if lastname belongs to unique roster."""
        return normalize_text(lastname) in [normalize_text(name) for name in self.unique_roster]
    
    def is_duplicate_lastname(self, lastname):
        """Check if lastname appears in duplicate roster."""
        for full_name in self.duplicate_roster:
            if ',' in full_name:
                stored_lastname = full_name.split(',')[0].strip()
                if normalize_text(stored_lastname) == normalize_text(lastname):
                    return True
        return False
    
    def get_students_with_lastname(self, lastname):
        """Get all students with specific lastname from duplicate roster."""
        students = []
        for full_name in self.duplicate_roster:
            if ',' in full_name:
                stored_lastname = full_name.split(',')[0].strip()
                if normalize_text(stored_lastname) == normalize_text(lastname):
                    students.append(full_name)
        return students
    
    def match_number_pronunciation(self, text):
        """Match number pronunciations - placeholder for now."""
        return None, 0

class DualRosterMatcher:
    """Enhanced name matching for dual roster system."""
    
    def __init__(self, roster_analyzer, pronunciation_trainer=None):
        self.roster_analyzer = roster_analyzer
        self.mappings = roster_analyzer.name_mappings
        self.pronunciation_trainer = pronunciation_trainer  # Now properly initialized

    def extract_names_from_speech(self, text):
        """Extract potential names from speech text with improved accuracy."""
        cleaned_text = normalize_text(text)
        words = cleaned_text.split()
        candidates = set()

        # Add single word candidates (for unique surnames)
        for w in words:
            if len(w) > 2:  # Only consider words longer than 2 characters
                candidates.add(w)

        # Add two-word combinations for full names
        if len(words) >= 2:
            for i in range(len(words)-1):
                candidates.add(f"{words[i]} {words[i+1]}")

        # Add the full text as a candidate
        candidates.add(cleaned_text)

        return sorted(candidates, key=len, reverse=True)  # Sort by length, longest first

    def match_name(self, text):
        """Enhanced name matching with audio-trained pronunciation data."""
        # Try audio-trained pronunciation matching first
        if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer and self.pronunciation_trainer.trained:
            name, score = self.pronunciation_trainer.match_pronunciation(text)
            if name and score >= 50:  # Lower threshold for audio-trained data
                return name, score, "audio_trained_match"
        
        # Fallback to existing fuzzy matching
        name_candidates = self.extract_names_from_speech(text)
        best_match = None
        best_score = 0
        match_type = None

        for candidate in name_candidates:
            candidate_norm = normalize_text(candidate)
            
            # Try exact matches first for unique last names
            if candidate_norm in self.mappings['unique_lastnames']:
                return self.mappings['unique_lastnames'][candidate_norm], 100, "exact_unique_lastname_match"
            
            # Try exact matches for full names in duplicate roster
            if candidate_norm in self.mappings['duplicate_fullnames']:
                return self.mappings['duplicate_fullnames'][candidate_norm], 100, "exact_duplicate_full_match"
            
            # Try exact matches for reversed full names in duplicate roster
            if candidate_norm in self.mappings['firstname_lastname']:
                return self.mappings['firstname_lastname'][candidate_norm], 100, "exact_firstname_lastname_match"

            # Special handling for compound names (Dela Cruz, etc.)
            if any(word in candidate_norm for word in ['dela', 'de la']):
                for stored_name in self.mappings['unique_lastnames'].values():
                    if any(word in stored_name.lower() for word in ['dela', 'de la']):
                        score = fuzz.token_set_ratio(candidate_norm, normalize_text(stored_name))
                        if score >= NAME_MATCH_THRESHOLD and score > best_score:
                            # Check for similar sounding names
                            if not any(self._is_similar_sounding(stored_name, other) 
                                     for other in self.mappings['unique_lastnames'].values() 
                                     if other != stored_name):
                                best_match = stored_name
                                best_score = score
                                match_type = "compound_match"
            
            # Regular fuzzy matching with sound similarity check for unique last names
            for stored_name in self.mappings['unique_lastnames'].values():
                # Skip if the current best match is already good enough
                if best_score >= 95:
                    continue
                    
                ratio_score = fuzz.ratio(candidate_norm, normalize_text(stored_name))
                token_score = fuzz.token_set_ratio(candidate_norm, normalize_text(stored_name))
                partial_score = fuzz.partial_ratio(candidate_norm, normalize_text(stored_name))
                
                score = max(ratio_score, token_score, partial_score)
                
                if score >= NAME_MATCH_THRESHOLD and score > best_score:
                    # Only accept if not too similar to other names
                    if not any(self._is_similar_sounding(stored_name, other) 
                             for other in self.mappings['unique_lastnames'].values() 
                             if other != stored_name):
                        best_match = stored_name
                        best_score = score
                        match_type = "fuzzy_match"
            
            # Also consider fuzzy matching for full names in duplicate roster
            for full_name in self.roster_analyzer.duplicate_roster:
                if ',' in full_name:
                    last_name, first_name = full_name.split(',', 1)
                    full_norm = normalize_text(f"{last_name.strip()} {first_name.strip()}")
                    
                    ratio_score = fuzz.ratio(candidate_norm, full_norm)
                    token_score = fuzz.token_set_ratio(candidate_norm, full_norm)
                    partial_score = fuzz.partial_ratio(candidate_norm, full_norm)
                    
                    score = max(ratio_score, token_score, partial_score)
                    
                    if score >= FULL_NAME_MATCH_THRESHOLD and score > best_score:
                        best_match = full_name
                        best_score = score
                        match_type = "duplicate_fuzzy_match"

        return best_match, best_score, match_type

    # Inside the DualRosterMatcher class, replace these existing methods:
    def _are_names_similar(self, name1, name2):
        """Check if names are too similar-sounding to risk confusion."""
        return self._is_similar_sounding(name1, name2)

    def is_similar_sounding(self, name1, name2):
        """Check if names sound too similar to avoid confusion."""
        return self._is_similar_sounding(name1, name2)
    
    def _is_similar_sounding(self, text1, text2):
        """Check if two texts sound too similar."""
        if not text1 or not text2:
            return False
            
        # Convert to simplified phonetic form
        t1 = text1.lower()
        t2 = text2.lower()
        
        # Common confusable sounds in Filipino
        replacements = {
            'v': 'b',
            'f': 'p',
            'z': 's',
            'x': 'ks',
            'ce': 'se',
            'ci': 'si',
            'ly': 'li',
            'll': 'l',
            'ñ': 'ny',
            'qu': 'k',
            'gue': 'ge',
            'gui': 'gi',
        }
        
        for old, new in replacements.items():
            t1 = t1.replace(old, new)
            t2 = t2.replace(old, new)
        
        # Return true if they sound too similar
        basic_ratio = fuzz.ratio(t1, t2)
        token_ratio = fuzz.token_set_ratio(t1, t2)
        return basic_ratio > 85 or token_ratio > 90  # High thresholds for sound similarity
    
    def get_phonetic_code(self, name):
        """Get phonetic code for name comparison."""
        if not name:
            return ""
        # Basic phonetic rules for Filipino names
        text = normalize_text(name)
        # Replace similar-sounding patterns
        replacements = [
            ('b', 'v'),
            ('f', 'p'),
            ('z', 's'),
            ('ce', 'se'),
            ('ci', 'si'),
            ('y', 'i'),
            ('ll', 'y'),
            ('gue', 'ge'),
            ('gui', 'gi'),
            ('qu', 'k'),
        ]
        for old, new in replacements:
            text = text.replace(old, new)
        return text

class SpeakerRecognizer:
    """Handles speaker diarization and teacher voice recognition."""
    
    def __init__(self):
        self.teacher_voice_embeddings = []
        self.teacher_voice_trained = False
        self.embedding_model = None
        
        if SPEAKER_DIARIZATION_AVAILABLE:
            self.setup_speaker_embedding_model()
            self.load_or_train_teacher_voice()
        else:
            print("⚠️ Speaker diarization disabled - all voices will be processed")
    
    def setup_speaker_embedding_model(self):
        """Initialize the speaker embedding model."""
        try:
            print("🔧 Setting up speaker embedding model...")
            self.embedding_model = PretrainedSpeakerEmbedding(
                "speechbrain/spkrec-ecapa-voxceleb",
                device=torch.device("cuda" if torch.cuda.is_available() else "cpu")
            )
            print("✅ Speaker embedding model ready")
        except Exception as e:
            print(f"❌ Failed to setup embedding model: {e}")
            try:
                from pyannote.audio import Model
                self.embedding_model = Model.from_pretrained("pyannote/embedding")
                print("✅ Alternative embedding model loaded")
            except Exception as e2:
                print(f"❌ Alternative model also failed: {e2}")
                self.embedding_model = None
    
    def load_or_train_teacher_voice(self):
        """Load existing teacher voice model or create new one."""
        if self.load_teacher_voice_model():
            print("✅ Teacher voice model loaded")
        else:
            print("🎤 No teacher voice model found. Please train one.")
    
    def train_teacher_voice(self, audio_samples_dir=None):
        """Train teacher voice recognition from audio samples."""
        if not SPEAKER_DIARIZATION_AVAILABLE or not self.embedding_model:
            print("❌ Speaker recognition not available")
            return False
        
        if audio_samples_dir is None:
            audio_samples_dir = TEACHER_VOICE_DIR
        
        if not os.path.exists(audio_samples_dir):
            print(f"❌ Teacher voice directory not found: {audio_samples_dir}")
            return False
        
        print("🎤 Training teacher voice recognition...")
        embeddings = []
        successful_files = []
        
        try:
            audio_files = [f for f in os.listdir(audio_samples_dir) 
                          if f.lower().endswith(('.wav', '.mp3', '.flac'))]
            
            if not audio_files:
                print("❌ No audio files found in directory")
                return False
            
            for audio_file in audio_files:
                audio_path = os.path.join(audio_samples_dir, audio_file)
                print(f"Processing teacher voice sample: {audio_file}")
                
                try:
                    embedding = self.extract_speaker_embedding(audio_path)
                    if embedding is not None:
                        embeddings.append(embedding)
                        successful_files.append(audio_file)
                        print(f"✅ Successfully processed: {audio_file}")
                    else:
                        print(f"⚠️ Failed to extract embedding from: {audio_file}")
                        
                except Exception as file_error:
                    print(f"❌ Error processing {audio_file}: {file_error}")
                    continue
            
            if embeddings:
                self.teacher_voice_embeddings = embeddings
                self.teacher_voice_trained = True
                self.save_teacher_voice_model()
                print(f"✅ Teacher voice trained with {len(embeddings)} samples")
                return True
            else:
                print("❌ No valid teacher voice embeddings could be extracted")
                return False
                
        except Exception as e:
            print(f"❌ Error in teacher voice training: {e}")
            traceback.print_exc()
            return False
    
    def extract_speaker_embedding(self, audio_path):
        """Extract speaker embedding from audio file."""
        try:
            if not self.embedding_model or not os.path.exists(audio_path):
                return None
            
            waveform, sample_rate = torchaudio.load(audio_path)
            
            # Check duration
            duration = waveform.shape[1] / sample_rate
            if duration < 0.5:
                print("⚠️ Audio too short")
                return None
            
            if duration > 30:
                waveform = waveform[:, :int(30 * sample_rate)]
            
            # Resample if necessary
            if sample_rate != SAMPLE_RATE:
                resampler = torchaudio.transforms.Resample(sample_rate, SAMPLE_RATE)
                waveform = resampler(waveform)
            
            # Convert to mono
            if waveform.shape[0] > 1:
                waveform = torch.mean(waveform, dim=0, keepdim=True)
            
            # Extract embedding based on model type
            if hasattr(self.embedding_model, '__call__'):
                # Speechbrain model
                if len(waveform.shape) == 1:
                    waveform_for_model = waveform.unsqueeze(0).unsqueeze(0)
                elif len(waveform.shape) == 2:
                    waveform_for_model = waveform.unsqueeze(0)
                else:
                    waveform_for_model = waveform
                
                embedding = self.embedding_model(waveform_for_model)
                if torch.is_tensor(embedding):
                    embedding = embedding.detach().cpu().numpy().flatten()
            else:
                # Pyannote model
                temp_file = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
                try:
                    torchaudio.save(temp_file.name, waveform, SAMPLE_RATE)
                    embedding = self.embedding_model(temp_file.name)
                    
                    if hasattr(embedding, 'data'):
                        embedding = embedding.data
                    if torch.is_tensor(embedding):
                        embedding = embedding.detach().cpu().numpy()
                    elif not isinstance(embedding, np.ndarray):
                        embedding = np.array(embedding)
                    
                    if len(embedding.shape) > 1:
                        embedding = np.mean(embedding, axis=0)
                        
                finally:
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
            
            # Ensure proper format
            if hasattr(embedding, 'shape') and len(embedding.shape) > 1:
                embedding = embedding.flatten()
            elif not isinstance(embedding, np.ndarray):
                embedding = np.array(embedding).flatten()
            
            if embedding is None or len(embedding) == 0:
                return None
            
            # Normalize
            embedding = embedding.astype(np.float32)
            norm = np.linalg.norm(embedding)
            if norm > 0:
                embedding = embedding / norm
            
            return embedding
            
        except Exception as e:
            print(f"❌ Error extracting speaker embedding: {e}")
            return None
    
    def save_teacher_voice_model(self):
        """Save teacher voice model to disk."""
        try:
            model_data = {
                'teacher_voice_embeddings': self.teacher_voice_embeddings,
                'teacher_voice_trained': self.teacher_voice_trained,
                'version': '2.0'
            }
            
            with open(TEACHER_VOICE_MODEL_PATH, 'wb') as f:
                pickle.dump(model_data, f)
            
            print(f"✅ Teacher voice model saved")
            
        except Exception as e:
            print(f"❌ Error saving teacher voice model: {e}")
    
    def load_teacher_voice_model(self):
        """Load teacher voice model from disk."""
        try:
            if os.path.exists(TEACHER_VOICE_MODEL_PATH):
                with open(TEACHER_VOICE_MODEL_PATH, 'rb') as f:
                    model_data = pickle.load(f)
                
                self.teacher_voice_embeddings = model_data.get('teacher_voice_embeddings', [])
                self.teacher_voice_trained = model_data.get('teacher_voice_trained', False)
                
                if self.teacher_voice_embeddings:
                    print(f"Loaded {len(self.teacher_voice_embeddings)} teacher voice embeddings")
                    return True
                
        except Exception as e:
            print(f"❌ Error loading teacher voice model: {e}")
        
        return False
    
    def is_teacher_voice(self, audio_data):
        """Check if the audio contains teacher's voice."""
        # Check if speaker recognition is disabled by user
        if hasattr(self, 'speaker_enabled') and not self.speaker_enabled.get():
            return True
            
        if not SPEAKER_DIARIZATION_AVAILABLE or not self.teacher_voice_trained or not self.embedding_model:
            return True  # If no speaker recognition, allow all voices
        
        try:
            temp_file = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
            
            try:
                audio_array = np.frombuffer(audio_data, dtype=np.int16)
                duration = len(audio_array) / SAMPLE_RATE
                
                if duration < MIN_SPEECH_DURATION:
                    return True
                
                with wave.open(temp_file.name, 'wb') as wav_file:
                    wav_file.setnchannels(CHANNELS)
                    wav_file.setsampwidth(2)
                    wav_file.setframerate(SAMPLE_RATE)
                    wav_file.writeframes(audio_data)
                
                current_embedding = self.extract_speaker_embedding(temp_file.name)
                
                if current_embedding is None:
                    print("Warning: Could not extract embedding, allowing voice")
                    return True
                
                max_similarity = 0
                similarities = []
                for i, teacher_embedding in enumerate(self.teacher_voice_embeddings):
                    try:
                        current_norm = current_embedding / (np.linalg.norm(current_embedding) + 1e-8)
                        teacher_norm = teacher_embedding / (np.linalg.norm(teacher_embedding) + 1e-8)
                        similarity = np.dot(current_norm, teacher_norm)
                        similarities.append(similarity)
                        max_similarity = max(max_similarity, similarity)
                    except Exception as e:
                        print(f"Error calculating similarity {i}: {e}")
                        continue
                
                # Debug output
                print(f"Voice similarity scores: {similarities}")
                print(f"Max similarity: {max_similarity:.3f}, Threshold: {SPEAKER_SIMILARITY_THRESHOLD}")
                print(f"Voice {'ACCEPTED' if max_similarity >= SPEAKER_SIMILARITY_THRESHOLD else 'REJECTED'}")
                
                return max_similarity >= SPEAKER_SIMILARITY_THRESHOLD
                
            finally:
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
            
        except Exception as e:
            print(f"Error in speaker recognition: {e}")
            return True
        
def save_text_log():
    """Save session summary."""
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary = [f"\n========== SESSION END: {ts} =========="]

        if 'app' in globals() and hasattr(app, 'total_attempts'):
            accuracy = (app.successful_matches / app.total_attempts * 100) if app.total_attempts > 0 else 0
            avg_speed = sum(app.processing_times) / len(app.processing_times) if app.processing_times else 0
            
            summary.extend([
                f"Total Attempts: {app.total_attempts}",
                f"Successful Matches: {app.successful_matches}",
                f"Accuracy: {accuracy:.1f}%",
                f"Average Speed: {avg_speed:.2f}s",
                f"Speaker Recognition: {'Active' if app.speaker_recognizer.teacher_voice_trained else 'Inactive'}"
            ])

        with open(LOG_TXT, "a", encoding="utf-8") as f:
            f.write("\n".join(summary) + "\n")

    except Exception as e:
        print(f"⚠️ Could not save session log: {e}")
        
class RecitationSystem:
    """Enhanced recitation system with dual roster support."""
    
    def __init__(self):
        print("🚀 Initializing Dual Roster Recitation System...")
        

        # Initialize components in correct order
        self.roster_analyzer = DualRosterAnalyzer(UNIQUE_LASTNAME_ROSTER, DUPLICATE_LASTNAME_ROSTER)
        self.pronunciation_trainer = PronunciationTrainer()  # Initialize first
        self.name_matcher = DualRosterMatcher(self.roster_analyzer, self.pronunciation_trainer)  # Pass pronunciation trainer
        self.speaker_recognizer = SpeakerRecognizer()
        
        self.setup_excel()
        self.setup_vosk()
        self.setup_audio()
        
        # Audio processing variables
        self.audio_buffer = []
        self.last_speech_time = time.time()
        self.processing_lock = threading.Lock()
        self.listening = False
        
        # Performance tracking
        self.total_attempts = 0
        self.successful_matches = 0
        self.needs_clarification = 0
        self.processing_times = []
        self.recognition_log = []
        self.speaker_rejections = 0
        
        # Initialize speaker enabled flag
        self.speaker_enabled = None  # Will be set in GUI setup

        self.setup_gui()
        print("✅ Dual roster recitation system initialized!")
    
    def setup_excel(self):
        """Setup Excel file for recitation tracking."""
        try:
            if not os.path.exists(OUTPUT_XLSX):
                wb = Workbook()
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(["Name", "Points", "Timestamp", "Total"])
                wb.save(OUTPUT_XLSX)
            
            self.wb = load_workbook(OUTPUT_XLSX)
            if SHEET_NAME in self.wb.sheetnames:
                self.ws = self.wb[SHEET_NAME]
            else:
                self.ws = self.wb.create_sheet(SHEET_NAME)
                self.ws.append(["Name", "Points", "Timestamp", "Total"])
            
            # Ensure Total column exists
            headers = [cell.value for cell in self.ws[1]]
            if len(headers) < 4 or headers[3] != "Total":
                self.ws.cell(row=1, column=4, value="Total")
                self.wb.save(OUTPUT_XLSX)
            
            print("✅ Excel setup complete")
            
        except Exception as e:
            print(f"❌ Excel setup error: {e}")
    
    def setup_vosk(self):
        """Setup Vosk model with dual roster keywords."""
        try:
            print(f"Loading Vosk model: {VOSK_MODEL_PATH}")
            start_time = time.time()
            
            if not os.path.exists(VOSK_MODEL_PATH):
                print(f"❌ Model not found: {VOSK_MODEL_PATH}")
                self.model = None
                self.recognizer = None
                self.model_info = "Model: NOT FOUND"
                return
            
            self.model = Model(VOSK_MODEL_PATH)
            self.recognizer = KaldiRecognizer(self.model, SAMPLE_RATE, json.dumps(KWS_LIST))
            self.recognizer.SetWords(False)
            
            load_time = time.time() - start_time
            print(f"✅ Vosk model loaded in {load_time:.1f}s with {len(KWS_LIST)} keywords")
            
            # Status info
            training_status = "Enhanced" if PRONUNCIATION_TRAINING_AVAILABLE else "Standard"
            speaker_status = "Speaker Recognition" if (SPEAKER_DIARIZATION_AVAILABLE and self.speaker_recognizer.teacher_voice_trained) else "All Voices"
            self.model_info = f"Vosk | {'Filipino' if USE_FILIPINO else 'English'} | {training_status} | {speaker_status} | Dual Roster"
            
        except Exception as e:
            print(f"❌ Vosk error: {e}")
            self.model = None
            self.recognizer = None
            self.model_info = "Model: LOAD FAILED"

    
    def setup_audio(self):
        """Setup audio stream."""
        try:
            self.p = pyaudio.PyAudio()
            self.stream = self.p.open(
                format=pyaudio.paInt16,
                channels=CHANNELS,
                rate=SAMPLE_RATE,
                input=True,
                frames_per_buffer=CHUNK
            )
            print("✅ Audio setup complete")
            
        except Exception as e:
            print(f"❌ Audio setup error: {e}")
            self.stream = None
    
    def setup_gui(self):
        """Setup GUI interface with speaker recognition controls."""
        try:
            self.root = tk.Tk()
            self.root.title("AI Recitation System with Speaker Recognition")
            self.root.geometry("1000x850")
            self.root.configure(bg='#f0f0f0')

            # Title
            title = tk.Label(self.root, text="AI RECITATION SYSTEM WITH SPEAKER RECOGNITION",
                           font=('Arial', 16, 'bold'), bg='#f0f0f0', fg='#2c3e50')
            title.pack(pady=10)

            # Model info
            self.model_label = tk.Label(self.root, text=getattr(self, 'model_info', 'Loading...'),
                                       font=('Arial', 9), bg='#f0f0f0', fg='#7f8c8d')
            self.model_label.pack()

            # Training status
            training_info = self.get_training_status_text()
            self.training_label = tk.Label(self.root, text=training_info,
                                          font=('Arial', 9, 'bold'), bg='#f0f0f0', fg='#27ae60')
            self.training_label.pack()

            # Speaker recognition status
            speaker_info = self.get_speaker_status_text()
            speaker_color = '#27ae60' if self.speaker_recognizer.teacher_voice_trained else '#e67e22'
            self.speaker_label = tk.Label(self.root, text=speaker_info,
                                         font=('Arial', 9, 'bold'), bg='#f0f0f0', fg=speaker_color)
            self.speaker_label.pack()

            # Speaker training controls - always show some controls
            speaker_controls = tk.Frame(self.root, bg='#f0f0f0')
            speaker_controls.pack(pady=5)

            # Always add the disable speaker recognition toggle
            self.speaker_enabled = tk.BooleanVar(value=False)  # Default to disabled for troubleshooting
            disable_speaker_cb = tk.Checkbutton(speaker_controls, text="Enable Speaker Recognition",
                                              variable=self.speaker_enabled,
                                              font=('Arial', 9), bg='#f0f0f0',
                                              command=self.on_speaker_toggle)
            disable_speaker_cb.pack(side='left', padx=5)

            if SPEAKER_DIARIZATION_AVAILABLE:
                train_voice_btn = tk.Button(speaker_controls, text="Train Teacher Voice",
                                           command=self.train_teacher_voice_dialog,
                                           font=('Arial', 9), bg='#9b59b6', fg='white')
                train_voice_btn.pack(side='left', padx=5)

                test_voice_btn = tk.Button(speaker_controls, text="Test Teacher Voice",
                                          command=self.test_teacher_voice,
                                          font=('Arial', 9), bg='#3498db', fg='white')
                test_voice_btn.pack(side='left', padx=5)
                
                debug_btn = tk.Button(speaker_controls, text="Debug Audio File",
                                     command=self.debug_audio_dialog,
                                     font=('Arial', 9), bg='#f39c12', fg='white')
                debug_btn.pack(side='left', padx=5)

                # Add threshold adjustment
                threshold_frame = tk.Frame(speaker_controls, bg='#f0f0f0')
                threshold_frame.pack(side='left', padx=5)
                
                tk.Label(threshold_frame, text="Threshold:", bg='#f0f0f0', font=('Arial', 8)).pack(side='left')
                self.threshold_var = tk.StringVar(value=str(SPEAKER_SIMILARITY_THRESHOLD))
                threshold_entry = tk.Entry(threshold_frame, textvariable=self.threshold_var, width=5, font=('Arial', 8))
                threshold_entry.pack(side='left')
                threshold_entry.bind('<Return>', self.update_threshold)
            else:
                tk.Label(speaker_controls, text="(Speaker libraries not installed)", 
                        font=('Arial', 8), fg='#7f8c8d', bg='#f0f0f0').pack(side='left', padx=10)

            # Status
            self.status_label = tk.Label(self.root, text="Ready to start",
                                        font=('Arial', 12), bg='#f0f0f0', fg='#27ae60')
            self.status_label.pack(pady=5)

            # Enhanced Stats with speaker info
            stats_frame = tk.Frame(self.root, bg='#f0f0f0')
            stats_frame.pack(pady=5)

            self.accuracy_label = tk.Label(stats_frame, text="Accuracy: 0% (0/0)",
                                          font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#e74c3c')
            self.accuracy_label.pack(side='left', padx=10)

            self.speed_label = tk.Label(stats_frame, text="Avg Speed: 0.0s",
                                       font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#3498db')
            self.speed_label.pack(side='left', padx=10)

            self.speaker_stats_label = tk.Label(stats_frame, text="Speaker Rejections: 0",
                                               font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#e67e22')
            self.speaker_stats_label.pack(side='left', padx=10)

            # Speech Recognition
            speech_frame = ttk.LabelFrame(self.root, text="Speech Recognition", padding=10)
            speech_frame.pack(pady=10, padx=20, fill='x')

            self.speech_button = tk.Button(speech_frame, text="START LISTENING",
                                          command=self.toggle_listening,
                                          font=('Arial', 12, 'bold'),
                                          bg='#3498db', fg='white', width=20)
            self.speech_button.pack(pady=5)

            # Enhanced transcription label with speaker info
            self.transcription_label = tk.Label(speech_frame, text="Say: 'Name + Points' (Teacher voice only)",
                                               font=('Arial', 10), wraplength=700)
            self.transcription_label.pack(pady=5)

            # Manual Entry
            manual_frame = ttk.LabelFrame(self.root, text="Manual Entry", padding=10)
            manual_frame.pack(pady=10, padx=20, fill='x')

            # Row 0: Student and Points
            tk.Label(manual_frame, text="Student:").grid(row=0, column=0, sticky='w', pady=2)
            self.student_var = tk.StringVar()
            self.student_combo = ttk.Combobox(manual_frame, textvariable=self.student_var,
                                             values=ROSTER, width=18, state='readonly')
            self.student_combo.grid(row=0, column=1, padx=5, pady=2)

            tk.Label(manual_frame, text="Points:").grid(row=0, column=2, sticky='w', padx=(20, 0), pady=2)
            self.points_var = tk.StringVar()
            self.points_entry = tk.Entry(manual_frame, textvariable=self.points_var, width=10)
            self.points_entry.grid(row=0, column=3, padx=5, pady=2)

            submit_btn = tk.Button(manual_frame, text="ADD POINTS", command=self.manual_submit,
                                  font=('Arial', 10, 'bold'), bg='#27ae60', fg='white')
            submit_btn.grid(row=0, column=4, padx=10, pady=2)

            # Row 1: Quick buttons
            quick_frame = tk.Frame(manual_frame)
            quick_frame.grid(row=1, column=0, columnspan=5, pady=10)

            tk.Label(quick_frame, text="Quick Add:").pack(side='left')
            for points in [1, 2, 3, 5]:
                btn = tk.Button(quick_frame, text=f"+{points}",
                               command=lambda p=points: self.quick_add_points(p),
                               font=('Arial', 9), bg='#f39c12', fg='white', width=5)
                btn.pack(side='left', padx=2)

            # Scores Table
            scores_frame = ttk.LabelFrame(self.root, text="Current Scores", padding=10)
            scores_frame.pack(pady=10, padx=20, fill='both', expand=True)

            self.tree = ttk.Treeview(scores_frame, columns=('Name', 'Latest', 'Total'), show='headings', height=6)
            self.tree.heading('Name', text='Student Name')
            self.tree.heading('Latest', text='Latest Points')
            self.tree.heading('Total', text='Total Points')

            self.tree.column('Name', width=260)
            self.tree.column('Latest', width=120, anchor='center')
            self.tree.column('Total', width=120, anchor='center')

            scrollbar = ttk.Scrollbar(scores_frame, orient='vertical', command=self.tree.yview)
            self.tree.configure(yscrollcommand=scrollbar.set)

            self.tree.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')

            # Enhanced Log with speaker recognition info
            log_frame = ttk.LabelFrame(self.root, text="AI Recognition & Speaker Log", padding=5)
            log_frame.pack(pady=10, padx=20, fill='both', expand=True)

            # Log controls
            log_controls = tk.Frame(log_frame)
            log_controls.pack(fill='x', pady=(0, 5))

            tk.Label(log_controls, text="Speech recognition with speaker filtering:",
                    font=('Arial', 9, 'bold')).pack(side='left')

            clear_btn = tk.Button(log_controls, text="Clear Log", command=self.clear_log,
                                 font=('Arial', 8), bg='#95a5a6', fg='white', width=10)
            clear_btn.pack(side='right')

            # Log text
            log_text_frame = tk.Frame(log_frame)
            log_text_frame.pack(fill='both', expand=True)

            self.log_text = tk.Text(log_text_frame, height=8, font=('Consolas', 9), wrap='word',
                                   bg='#2c3e50', fg='#ecf0f1', insertbackground='white')
            log_scrollbar = ttk.Scrollbar(log_text_frame, orient='vertical', command=self.log_text.yview)
            self.log_text.configure(yscrollcommand=log_scrollbar.set)

            self.log_text.pack(side='left', fill='both', expand=True)
            log_scrollbar.pack(side='right', fill='y')

            # Initial log messages
            self.add_to_log("=== AI RECITATION SYSTEM WITH SPEAKER RECOGNITION ===")
            self.add_to_log(f"Speech Model: Vosk ({'Filipino' if USE_FILIPINO else 'English'})")
            
            if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
                sample_count = sum(len(data['transcripts']) for data in self.pronunciation_trainer.pronunciation_data.values())
                self.add_to_log(f"Enhanced matching: {sample_count} pronunciation samples")
            
            if SPEAKER_DIARIZATION_AVAILABLE:
                if self.speaker_recognizer.teacher_voice_trained:
                    self.add_to_log(f"Speaker Recognition: ACTIVE - Teacher voice trained")
                else:
                    self.add_to_log(f"Speaker Recognition: Train teacher voice for filtering")
            else:
                self.add_to_log("Speaker Recognition: DISABLED (missing libraries)")
            
            self.add_to_log("Click 'START LISTENING' to begin recognition...")

            # Bind Enter key
            self.points_entry.bind('<Return>', lambda e: self.manual_submit())

            # Initial data load
            self.refresh_scores()
            print("✅ GUI setup complete")
            
        except Exception as e:
            print(f"❌ GUI setup error: {e}")
            traceback.print_exc()
        

    def get_training_status_text(self):
        """Get pronunciation training status text."""
        if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
            sample_count = sum(len(data['transcripts']) for data in self.pronunciation_trainer.pronunciation_data.values())
            return f"Pronunciation Training: ACTIVE ({sample_count} samples)"
        else:
            return "Pronunciation Training: INACTIVE"

    def get_speaker_status_text(self):
        """Get speaker recognition status text."""
        if SPEAKER_DIARIZATION_AVAILABLE:
            if self.speaker_recognizer.teacher_voice_trained:
                embedding_count = len(self.speaker_recognizer.teacher_voice_embeddings)
                return f"Speaker Recognition: ACTIVE ({embedding_count} voice samples)"
            else:
                return "Speaker Recognition: READY (needs teacher voice training)"
        else:
            return "Speaker Recognition: UNAVAILABLE (missing libraries)"

    def on_speaker_toggle(self):
        """Handle speaker recognition enable/disable toggle."""
        if self.speaker_enabled.get():
            self.add_to_log("Speaker recognition ENABLED")
            if not SPEAKER_DIARIZATION_AVAILABLE:
                self.add_to_log("WARNING: Speaker libraries not available!")
                self.speaker_enabled.set(False)
            elif not self.speaker_recognizer.teacher_voice_trained:
                self.add_to_log("WARNING: Teacher voice not trained yet!")
        else:
            self.add_to_log("Speaker recognition DISABLED - all voices will be processed")

    def update_threshold(self, event=None):
        """Update the speaker similarity threshold."""
        try:
            new_threshold = float(self.threshold_var.get())
            if 0.0 <= new_threshold <= 1.0:
                global SPEAKER_SIMILARITY_THRESHOLD
                SPEAKER_SIMILARITY_THRESHOLD = new_threshold
                self.add_to_log(f"Speaker threshold updated to: {new_threshold}")
            else:
                messagebox.showwarning("Invalid Threshold", "Threshold must be between 0.0 and 1.0")
                self.threshold_var.set(str(SPEAKER_SIMILARITY_THRESHOLD))
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid number")
            self.threshold_var.set(str(SPEAKER_SIMILARITY_THRESHOLD))

    def debug_audio_dialog(self):
        """Dialog to debug audio files."""
        file_path = filedialog.askopenfilename(
            title="Select Audio File to Debug",
            filetypes=[("Audio Files", "*.wav *.mp3 *.flac"), ("All Files", "*.*")]
        )
        
        if file_path:
            self.add_to_log(f"Debugging audio file: {os.path.basename(file_path)}")
            debug_result = debug_audio_file(file_path)
            
            if debug_result:
                self.add_to_log("Audio file appears to be valid")
            else:
                self.add_to_log("Audio file has issues - check console for details")

    def train_teacher_voice_dialog(self):
        """Open dialog for training teacher voice - ENHANCED WITH DEBUG."""
        if not SPEAKER_DIARIZATION_AVAILABLE:
            messagebox.showerror("Error", "Speaker recognition libraries not available.\nInstall: pip install pyannote-audio torch torchaudio speechbrain")
            return

        # Create teacher voice directory if it doesn't exist
        os.makedirs(TEACHER_VOICE_DIR, exist_ok=True)

        # Ask user to select teacher voice samples
        result = messagebox.askyesno(
            "Train Teacher Voice",
            f"To train teacher voice recognition:\n\n"
            f"1. Record 3-5 audio samples of the teacher speaking\n"
            f"2. Save them as .wav files in:\n   {TEACHER_VOICE_DIR}\n"
            f"3. Each sample should be 2-5 seconds long\n"
            f"4. Use clear speech with minimal background noise\n\n"
            f"Would you like to select existing audio files now?"
        )

        if result:
            # Let user select multiple audio files
            file_paths = filedialog.askopenfilenames(
                title="Select Teacher Voice Samples",
                filetypes=[("Audio Files", "*.wav *.mp3 *.flac"), ("All Files", "*.*")]
            )

            if file_paths:
                self.add_to_log(f"Selected {len(file_paths)} files for training")
                
                # Debug each file first
                valid_files = []
                for i, file_path in enumerate(file_paths):
                    self.add_to_log(f"Debugging file {i+1}: {os.path.basename(file_path)}")
                    
                    if debug_audio_file(file_path):
                        valid_files.append(file_path)
                        self.add_to_log(f"File {i+1} is valid")
                    else:
                        self.add_to_log(f"File {i+1} has issues - skipping")

                if not valid_files:
                    messagebox.showerror("Error", "No valid audio files found. Check the debug output in the log.")
                    return

                self.add_to_log(f"{len(valid_files)} out of {len(file_paths)} files are valid")

                # Copy valid files to teacher voice directory
                copied_count = 0
                for i, file_path in enumerate(valid_files):
                    try:
                        filename = f"teacher_sample_{i+1}_{int(time.time())}.wav"
                        dest_path = os.path.join(TEACHER_VOICE_DIR, filename)
                        
                        # Convert to proper WAV format
                        waveform, sample_rate = torchaudio.load(file_path)
                        
                        # Ensure mono and proper sample rate
                        if waveform.shape[0] > 1:
                            waveform = torch.mean(waveform, dim=0, keepdim=True)
                        
                        if sample_rate != SAMPLE_RATE:
                            resampler = torchaudio.transforms.Resample(sample_rate, SAMPLE_RATE)
                            waveform = resampler(waveform)
                        
                        # Save converted file
                        torchaudio.save(dest_path, waveform, SAMPLE_RATE)
                        copied_count += 1
                        self.add_to_log(f"Converted and saved: {filename}")
                        
                    except Exception as e:
                        self.add_to_log(f"Error processing {file_path}: {e}")
                        print(f"Error copying {file_path}: {e}")

                if copied_count > 0:
                    self.add_to_log(f"Successfully processed {copied_count} teacher voice samples")
                    
                    # Train the teacher voice model
                    self.add_to_log("Training teacher voice recognition...")
                    success = self.speaker_recognizer.train_teacher_voice()
                    
                    if success:
                        self.add_to_log("Teacher voice training completed!")
                        # Update GUI status
                        self.speaker_label.config(
                            text=self.get_speaker_status_text(),
                            fg='#27ae60'
                        )
                        messagebox.showinfo("Success", "Teacher voice trained successfully!")
                    else:
                        self.add_to_log("Teacher voice training failed")
                        messagebox.showerror("Error", "Failed to train teacher voice. Check audio files and console output.")
                else:
                    messagebox.showerror("Error", "No valid audio files could be processed.")
        else:
            messagebox.showinfo(
                "Manual Training",
                f"To train manually:\n\n"
                f"1. Record teacher voice samples\n"
                f"2. Save as .wav files in: {TEACHER_VOICE_DIR}\n"
                f"3. Click 'Train Teacher Voice' button again"
            )

    def test_teacher_voice(self):
        """Test teacher voice recognition with current audio."""
        if not SPEAKER_DIARIZATION_AVAILABLE or not self.speaker_recognizer.teacher_voice_trained:
            messagebox.showwarning("Warning", "Teacher voice not trained yet. Please train first.")
            return

        # Record a short audio sample for testing
        messagebox.showinfo("Voice Test", "Click OK, then speak for 3 seconds to test teacher voice recognition.")
        
        try:
            # Record 3 seconds of audio
            test_duration = 3
            test_chunks = int(SAMPLE_RATE * test_duration / CHUNK)
            test_audio = []
            
            for _ in range(test_chunks):
                data = self.stream.read(CHUNK, exception_on_overflow=False)
                test_audio.append(data)
            
            # Test if it's teacher voice
            test_audio_data = b''.join(test_audio)
            is_teacher = self.speaker_recognizer.is_teacher_voice(test_audio_data)
            
            if is_teacher:
                messagebox.showinfo("Test Result", "TEACHER VOICE DETECTED!\nThis voice will be processed.")
                self.add_to_log("Voice test: TEACHER VOICE confirmed")
            else:
                messagebox.showinfo("Test Result", "NOT TEACHER VOICE\nThis voice will be ignored.")
                self.add_to_log("Voice test: NOT teacher voice")
                
        except Exception as e:
            messagebox.showerror("Error", f"Voice test failed: {e}")
            self.add_to_log(f"Voice test error: {e}")


    def toggle_listening(self):
        """Toggle speech recognition."""
        if not self.listening:
            if self.model and self.stream:
                self.listening = True
                self.speech_button.config(text="STOP LISTENING", bg='#e74c3c')
                if self.speaker_recognizer.teacher_voice_trained:
                    self.status_label.config(text="Listening for TEACHER voice only...", fg='#e74c3c')
                else:
                    self.status_label.config(text="Listening (all voices)...", fg='#e74c3c')
                self.add_to_log("STARTED listening with speaker filtering")
                threading.Thread(target=self.audio_loop, daemon=True).start()
            else:
                messagebox.showerror("Error", "Speech recognition not available.")
                self.add_to_log("Speech recognition unavailable")
        else:
            self.listening = False
            self.speech_button.config(text="START LISTENING", bg='#3498db')
            self.status_label.config(text="Stopped listening", fg='#f39c12')
            self.add_to_log("STOPPED listening")

    def audio_loop(self):
        """Main audio processing loop with speaker filtering."""
        while self.listening and self.stream:
            try:
                data = self.stream.read(CHUNK, exception_on_overflow=False)
                current_time = time.time()

                if not self.is_silence(data):
                    self.audio_buffer.append(data)
                    self.last_speech_time = current_time
                else:
                    if self.audio_buffer and (current_time - self.last_speech_time) > SILENCE_DURATION:
                        self.process_audio_buffer()

                # Prevent buffer overflow
                max_chunks = int(SAMPLE_RATE * BUFFER_SECONDS / CHUNK)
                if len(self.audio_buffer) > max_chunks:
                    self.process_audio_buffer()

            except Exception as e:
                print(f"Audio loop error: {e}")
                self.root.after(0, lambda: self.add_to_log(f"[ERROR] Audio error: {e}"))
                break

    def is_silence(self, data: bytes) -> bool:
        """Check if audio data is silence using RMS threshold."""
        try:
            audio = np.frombuffer(data, dtype=np.int16).astype(np.float32)
            if audio.size == 0:
                return True
            rms = np.sqrt(np.mean(np.square(audio))) + 1e-8
            return rms < 300.0
        except Exception:
            return False

    def process_audio_buffer(self):
        """Process accumulated audio with speaker filtering and Vosk."""
        with self.processing_lock:
            if not self.audio_buffer or not self.model:
                self.audio_buffer = []
                self.root.after(0, lambda: self.status_label.config(text="Listening for teacher voice...", fg='#e74c3c'))
                return

            try:
                start_time = time.time()
                
                # Join audio data
                audio_data = b''.join(self.audio_buffer)
                
                # Check if this is teacher's voice (speaker filtering)
                # Only do speaker filtering if enabled AND available AND trained
                should_filter_speaker = (
                    hasattr(self, 'speaker_enabled') and 
                    self.speaker_enabled.get() and
                    SPEAKER_DIARIZATION_AVAILABLE and 
                    self.speaker_recognizer.teacher_voice_trained
                )
                
                if should_filter_speaker:
                    self.root.after(0, lambda: self.status_label.config(text="Checking speaker...", fg='#f39c12'))
                    
                    is_teacher = self.speaker_recognizer.is_teacher_voice(audio_data)
                    
                    if not is_teacher:
                        # Reject non-teacher voice
                        self.speaker_rejections += 1
                        self.update_speaker_stats()
                        self.root.after(0, lambda: self.add_to_log("REJECTED: Non-teacher voice detected"))
                        self.root.after(0, lambda: self.status_label.config(text="Listening for teacher voice...", fg='#e74c3c'))
                        self.audio_buffer = [] # Clear buffer on rejection
                        return
                
                # If speaker filtering is disabled, or teacher voice is detected, proceed with Vosk processing.
                self.root.after(0, lambda: self.status_label.config(text="Processing teacher voice...", fg='#27ae60'))

                # Create fresh recognizer with grammar and alternatives
                recognizer = KaldiRecognizer(self.model, SAMPLE_RATE, json.dumps(KWS_LIST))
                recognizer.SetWords(False)
                try:
                    recognizer.SetMaxAlternatives(5)
                except Exception:
                    pass

                # Process with Vosk
                accepted = recognizer.AcceptWaveform(audio_data)
                if accepted:
                    result_text = recognizer.Result()
                else:
                    result_text = recognizer.FinalResult()

                # Parse result with alternatives and try best-of for matching
                try:
                    result_json = json.loads(result_text)
                except Exception:
                    result_json = {}

                text_main = (result_json.get("text") or "").strip()
                alts = result_json.get("alternatives", []) or []
                alt_texts = [a.get("text", "").strip() for a in alts if a.get("text")]
                candidates_to_try = [t for t in [text_main] + alt_texts if t]

                processing_time = time.time() - start_time
                self.processing_times.append(processing_time)
                timestamp = datetime.now().strftime("%H:%M:%S")

                if candidates_to_try:
                    self.total_attempts += 1
                    display_text = text_main if text_main else "[no speech detected]"
                    self.root.after(0, lambda: self.transcription_label.config(text=f"Teacher said: '{display_text}'"))

                    method = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
                    speaker_info = "Teacher" if self.speaker_recognizer.teacher_voice_trained else "Voice"
                    self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] {speaker_info} | {method}: '{text_main}' | {processing_time:.2f}s"))

                    best_name, best_score, best_text, best_points = None, 0, "", None
                    
                    # Iterate through candidates for the best match, prioritizing alternatives with points
                    for t in candidates_to_try:
                        pts_try = self.extract_points(t)
                        name_try, score_try, match_kind = self.name_matcher.match_name(t)
                        
                        # Heuristic: a combination of name match score and a bonus if points are detected
                        effective_score = (score_try or 0) + (10 if pts_try is not None else 0) 
                        
                        if effective_score > best_score:
                            best_name, best_score, best_text, best_points = name_try, effective_score, t, pts_try

                    if best_name and best_score >= 50:  # Lower threshold
                        if best_points is not None:
                            # Both name and points found - log it
                            self.log_points(best_name, best_points, processing_time, match_score=best_score, method=method)
                            self.successful_matches += 1
                            self.root.after(0, self.update_performance_stats)  # ADD THIS
                            self.root.after(0, lambda: self.add_to_log(f"SUCCESS - Points logged: {best_name} +{best_points}", text_main))
                            self.root.after(0, lambda: self.status_label.config(text=f"{best_name} +{best_points} pts! (Teacher: {method}: {best_score:.0f}%)", fg='#27ae60'))
                        else:
                            # Name found but no points - ask for clarification
                            self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] PARTIAL - Name '{best_name}' found but no points detected in: '{text_main}'"))
                            self.root.after(0, lambda: self.status_label.config(text=f"Found {best_name} - please specify points", fg='#f39c12'))
                    else:
                        # No confident match
                        self.root.after(0, lambda: self.add_to_log(f"[{timestamp}] FAILED - No confident match found in: '{text_main}'"))
                        self.root.after(0, lambda: self.status_label.config(text="Please speak more clearly", fg='#e74c3c'))

                # Clear buffer
                self.audio_buffer = []

            except Exception as e:
                print(f"Processing error: {e}")
                traceback.print_exc()
                self.root.after(0, lambda: self.add_to_log(f"[ERROR] Processing failed: {str(e)}"))
                self.audio_buffer = [] # Always clear buffer on error
                self.root.after(0, lambda: self.status_label.config(text="Listening for teacher voice...", fg='#e74c3c'))

    def process_speech_text(self, text, processing_time=0.0):
        """Process speech text with more lenient confidence requirements."""
        points = self.extract_points(text)
        name, score, match_type = self.name_matcher.match_name(text)
        
        # Be more lenient with confidence threshold
        if name and score >= NAME_MATCH_THRESHOLD:
            if points is not None:
                method = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
                self.log_points(name, points, processing_time, match_score=score, method=method)
                self.status_label.config(text=f"Logged: {name} +{points} pts ({score:.1f}% confidence)", fg='#27ae60')
                return True
            else:
                self.add_to_log("Name recognized but no points detected")
                return False
        else:
            self.add_to_log(f"Low confidence in name recognition ({score:.1f}%)")
            return False

    def extract_points(self, text):
        """Extract points from text using enhanced number recognition."""
        if PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained:
            number, score = self.pronunciation_trainer.match_number_pronunciation(text)
            if number and score >= 60: 
                print(f"Enhanced number match: '{text}' -> {number} (confidence: {score:.1f}%)")
                return number
        
        number_words = {
            "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
            "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
            "isa": 1, "dalawa": 2, "tatlo": 3, "apat": 4, "lima": 5,
            "anim": 6, "pito": 7, "walo": 8, "siyam": 9, "sampu": 10,
            "uno": 1, "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5,
            "seis": 6, "siete": 7, "ocho": 8, "nueve": 9, "diez": 10
        }
        
        txt = (text or "").lower()
        
        m = re.search(r"\b(\d{1,2})\b", txt)
        if m:
            try:
                val = int(m.group(1))
                if 1 <= val <= 10:  # Valid points range
                    return val
            except Exception:
                pass
        
        for word, value in number_words.items():
            if re.search(rf"\b{re.escape(word)}\b", txt):
                return value
        
        m2 = re.search(r"\b(?:plus|add|dagdag|bigay)\s+(\d{1,2})\b", txt)
        if m2:
            try:
                val = int(m2.group(1))
                if 1 <= val <= 10:
                    return val
            except Exception:
                pass
        
        # Pattern 4: "plus/add" + number word
        for word, value in number_words.items():
            if re.search(rf"\b(?:plus|add|dagdag|bigay)\s+{re.escape(word)}\b", txt):
                return value
        
        # Pattern 5: number + "points" keyword
        m3 = re.search(r"\b(\d{1,2})\s*(?:pts|points?|puntos)\b", txt)
        if m3:
            try:
                val = int(m3.group(1))
                if 1 <= val <= 10:
                    return val
            except Exception:
                pass
        
        # Pattern 6: number word + "points" keyword
        for word, value in number_words.items():
            if re.search(rf"\b{re.escape(word)}\s*(?:pts|points?|puntos)\b", txt):
                return value
        
        return None

    def manual_submit(self):
        """Handle manual point submission."""
        student = self.student_var.get()
        points_str = self.points_var.get()

        if not student:
            messagebox.showwarning("Warning", "Please select a student")
            return

        try:
            points = int(points_str)
            if points <= 0:
                raise ValueError()
        except Exception:
            messagebox.showwarning("Warning", "Please enter a valid positive number for points")
            return

        self.log_points(student, points, processing_time=0.0, manual=True)
        self.points_var.set("")
        self.status_label.config(text=f"{student} +{points} points (manual)", fg='#27ae60')

    def quick_add_points(self, points):
        """Quick add points to selected student."""
        student = self.student_var.get()
        if not student:
            messagebox.showwarning("Warning", "Please select a student first")
            return

        self.log_points(student, points, processing_time=0.0, manual=True)
        self.status_label.config(text=f"{student} +{points} points (quick)", fg='#27ae60')

    def log_points(self, name, points, processing_time=0.0, match_score=None, manual=False, method="Standard"):
        """Log points to Excel and text log."""
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        found = False

        # Update Excel
        for row in range(2, self.ws.max_row + 1):
            cell_name = self.ws.cell(row=row, column=1).value
            if cell_name and cell_name.lower() == name.lower():
                old_points = self.ws.cell(row=row, column=2).value
                if old_points:
                    self.ws.cell(row=row, column=2, value=f"{old_points}+{points}")
                else:
                    self.ws.cell(row=row, column=2, value=str(points))

                self.ws.cell(row=row, column=3, value=ts)

                current_total = self.ws.cell(row=row, column=4).value or 0
                self.ws.cell(row=row, column=4, value=current_total + points)

                found = True
                break

        if not found:
            next_row = self.ws.max_row + 1
            self.ws.cell(row=next_row, column=1, value=name)
            self.ws.cell(row=next_row, column=2, value=str(points))
            self.ws.cell(row=next_row, column=3, value=ts)
            self.ws.cell(row=next_row, column=4, value=points)

        # Save Excel
        try:
            self.wb.save(OUTPUT_XLSX)
        except Exception as e:
            print(f"Failed to save Excel: {e}")

        # Enhanced text logging with speaker info
        total_now = self.get_total(name)
        speed_part = f" | Speed: {processing_time:.2f}s" if processing_time > 0 else ""
        score_part = f" | {method}Score: {match_score:.0f}%" if match_score is not None else ""
        manual_part = " (manual)" if manual else " (teacher voice)" if not manual else ""
        text_log_line = f"{ts} | {name} +{points}{manual_part} | Total: {total_now}{score_part}{speed_part}"

        try:
            with open(LOG_TXT, "a", encoding="utf-8") as f:
                f.write(text_log_line + "\n")
        except Exception as e:
            print(f"Failed to write text log: {e}")

        self.recognition_log.append(text_log_line)
        self.root.after(0, self.refresh_scores)
        self.root.after(0, self.update_performance_stats)


    def get_total(self, name):
        """Return numeric total for a student."""
        for row in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=1).value
            if val and val.lower() == name.lower():
                return self.ws.cell(row=row, column=4).value or 0
        return 0

    def update_performance_stats(self):
        """Update performance statistics display."""
        if self.total_attempts > 0:
            accuracy = (self.successful_matches / self.total_attempts) * 100
            accuracy_color = '#27ae60' if accuracy >= 70 else '#f39c12' if accuracy >= 50 else '#e74c3c'
            self.accuracy_label.config(
                text=f"Accuracy: {accuracy:.0f}% ({self.successful_matches}/{self.total_attempts})",
                fg=accuracy_color
            )
        else:
            self.accuracy_label.config(text="Accuracy: 0% (0/0)", fg='#e74c3c')

        if self.processing_times:
            avg_speed = sum(self.processing_times) / len(self.processing_times)
            speed_color = '#27ae60' if avg_speed <= 2.0 else '#f39c12' if avg_speed <= 4.0 else '#e74c3c'
            self.speed_label.config(text=f"Avg Speed: {avg_speed:.2f}s", fg=speed_color)
        else:
            self.speed_label.config(text="Avg Speed: 0.00s", fg='#3498db')

    def update_speaker_stats(self):
        """Update speaker rejection statistics."""
        self.speaker_stats_label.config(text=f"Speaker Rejections: {self.speaker_rejections}")

    def add_to_log(self, message, original_speech=None):
        """Add message to recognition log widget with enhanced color coding."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        # Add teacher's original speech in quotes if provided
        if original_speech:
            formatted_message = f"[{timestamp}] Teacher said: \"{original_speech}\" | {message}"
        else:
            formatted_message = f"[{timestamp}] {message}"
        
        self.log_text.insert(tk.END, formatted_message + "\n")
        self.log_text.see(tk.END)

        # Enhanced color coding for speaker recognition
        try:
            if "SUCCESS" in message:
                self.log_text.tag_add("success", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("success", foreground='#2ecc71')
            elif "REJECTED" in message:
                self.log_text.tag_add("rejected", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("rejected", foreground='#e67e22')
            elif "Teacher" in message:
                self.log_text.tag_add("teacher", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("teacher", foreground='#9b59b6')
            elif "ERROR" in message or "FAILED" in message:
                self.log_text.tag_add("error", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("error", foreground='#e74c3c')
            elif "Enhanced:" in message:
                self.log_text.tag_add("enhanced", f"{float(self.log_text.index('end'))-2}.0", tk.END)
                self.log_text.tag_config("enhanced", foreground='#3498db')
        except Exception:
            pass

        try:
            lines = int(self.log_text.index('end-1c').split('.')[0])
            if lines > 200:
                self.log_text.delete("1.0", f"{lines-200}.0")
        except Exception:
            pass

    def clear_log(self):
        """Clear the recognition log and reset stats."""
        self.log_text.delete("1.0", tk.END)
        self.total_attempts = 0
        self.successful_matches = 0
        self.processing_times = []
        self.recognition_log = []
        self.speaker_rejections = 0
        self.update_performance_stats()
        self.update_speaker_stats()
        self.add_to_log("=== LOG CLEARED - STATS RESET ===")
        status = "Enhanced" if (PRONUNCIATION_TRAINING_AVAILABLE and self.pronunciation_trainer.trained) else "Standard"
        speaker_status = "with Speaker Recognition" if self.speaker_recognizer.teacher_voice_trained else "All Voices"
        self.add_to_log(f"System ready: {status} matching {speaker_status}")

    def refresh_scores(self):
        """Refresh the Treeview with current Excel scores."""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Add current scores
        for row in range(2, self.ws.max_row + 1):
            name = self.ws.cell(row=row, column=1).value
            history = self.ws.cell(row=row, column=2).value or ""
            total = self.ws.cell(row=row, column=4).value or 0

            # Get latest points from history
            latest = 0
            if isinstance(history, str) and history.strip():
                try:
                    last_part = history.split("+")[-1]
                    latest = int(last_part)
                except Exception:
                    try:
                        latest = int(history)
                    except Exception:
                        latest = 0
            elif isinstance(history, (int, float)):
                latest = int(history)

            if name:
                self.tree.insert('', 'end', values=(name, latest, total))

    def run(self):
        """Start the GUI mainloop."""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def on_closing(self):
        """Cleanup on window close."""
        self.listening = False
        try:
            if self.stream:
                self.stream.stop_stream()
                self.stream.close()
            if hasattr(self, 'p'):
                self.p.terminate()
        except Exception:
            pass

        try:
            self.wb.save(OUTPUT_XLSX)
        except Exception:
            pass

        save_text_log()
        try:
            self.root.destroy()
        except Exception:
            pass


def debug_audio_file(audio_path):
    """Debug function to analyze audio file properties"""
    print(f"\n=== DEBUGGING AUDIO FILE: {audio_path} ===")
    
    # Check file existence and size
    if not os.path.exists(audio_path):
        print("File does not exist")
        return False
    
    file_size = os.path.getsize(audio_path)
    print(f"File size: {file_size} bytes ({file_size/1024:.1f} KB)")
    
    if file_size == 0:
        print("File is empty")
        return False
    
    # Try to load with torchaudio
    try:
        waveform, sample_rate = torchaudio.load(audio_path)
        duration = waveform.shape[1] / sample_rate
        channels = waveform.shape[0]
        
        print(f"Successfully loaded audio:")
        print(f"  - Duration: {duration:.2f} seconds")
        print(f"  - Sample rate: {sample_rate} Hz")
        print(f"  - Channels: {channels}")
        print(f"  - Shape: {waveform.shape}")
        
        # Check audio content
        max_val = torch.max(torch.abs(waveform)).item()
        mean_val = torch.mean(torch.abs(waveform)).item()
        
        print(f"  - Max amplitude: {max_val:.4f}")
        print(f"  - Mean amplitude: {mean_val:.4f}")
        
        if max_val < 0.001:
            print("Audio signal very weak (might be silence)")
        
        if duration < 1.0:
            print("Audio shorter than 1 second")
        elif duration > 10.0:
            print("Audio longer than 10 seconds")
        else:
            print("Audio duration is appropriate")
            
        return True
        
    except Exception as e:
        print(f"Failed to load audio: {e}")
        return False


if __name__ == "__main__":
    print("Starting Enhanced Recitation System with Speaker Recognition...")
    try:
        app = RecitationSystem()
        atexit.register(save_text_log)
        app.run()
    except Exception as e:
        print(f"Failed to start application: {e}")
        traceback.print_exc()
        input("Press Enter to exit...")
    print("Application closed.")
